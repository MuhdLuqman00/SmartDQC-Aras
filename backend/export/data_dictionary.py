"""Client-friendly Data Dictionary exports (Excel + PDF).

The /data-dictionary endpoint returns technical JSON; non-technical users
need a readable spreadsheet or printable document. Both renderers flatten the
source + derived field definitions into plain rows:
    Field | Category | Type | Description
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .report_template_spec import BRAND_NAVY, BRAND_BG

# Bahasa Melayu type labels — BM is the default locale.
_TYPE_BM = {
    "identifier":  "Pengenalan",
    "categorical": "Kategori",
    "numerical":   "Nombor",
    "boolean":     "Benar/Palsu",
    "date":        "Tarikh",
    "text":        "Teks",
}

HEADERS = ["Medan", "Kategori", "Jenis", "Keterangan"]


def _flatten(source_fields: dict, derived_fields: dict) -> list[list[str]]:
    """Return ordered rows: source fields first, then derived fields."""
    rows: list[list[str]] = []
    for name, meta in (source_fields or {}).items():
        t = (meta or {}).get("type", "")
        rows.append([name, "Medan Sumber", _TYPE_BM.get(t, t),
                     (meta or {}).get("description", "")])
    for name, meta in (derived_fields or {}).items():
        t = (meta or {}).get("type", "")
        rows.append([name, "Medan Terbitan", _TYPE_BM.get(t, t),
                     (meta or {}).get("description", "")])
    return rows


def to_excel(source_fields: dict, derived_fields: dict) -> bytes:
    rows = _flatten(source_fields, derived_fields)
    wb = Workbook()
    ws = wb.active
    ws.title = "Kamus Data"

    navy = BRAND_NAVY.lstrip("#")
    header_fill = PatternFill("solid", fgColor=navy)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    band_fill = PatternFill("solid", fgColor=BRAND_BG.lstrip("#"))
    thin = Side(style="thin", color="D9DEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:D1")
    ws["A1"] = "SmartDQC — Kamus Data / Data Dictionary"
    ws["A1"].font = Font(bold=True, size=14, color=navy)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    # Header row
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border

    # Data rows (zebra banding for readability)
    for i, row in enumerate(rows):
        r = i + 3
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=(col == 4))
            c.border = border
            if i % 2 == 1:
                c.fill = band_fill

    widths = [26, 16, 14, 70]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def to_pdf(source_fields: dict, derived_fields: dict) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    rows = _flatten(source_fields, derived_fields)
    navy = colors.HexColor(BRAND_NAVY)
    band = colors.HexColor(BRAND_BG)

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=8, leading=10)
    head_style = ParagraphStyle("cellh", parent=styles["Normal"], fontSize=8.5,
                                leading=10, textColor=colors.white, fontName="Helvetica-Bold")
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=16, textColor=navy)

    data = [[Paragraph(h, head_style) for h in HEADERS]]
    for row in rows:
        data.append([Paragraph(str(v), cell_style) for v in row])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title="SmartDQC Kamus Data",
    )
    table = Table(data, colWidths=[4.0 * cm, 2.8 * cm, 2.3 * cm, 8.0 * cm], repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9DEE8")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ])
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.add("BACKGROUND", (0, i), (-1, i), band)
    table.setStyle(style)

    elems = [
        Paragraph("SmartDQC — Kamus Data / Data Dictionary", title_style),
        Spacer(1, 0.3 * cm),
        table,
    ]
    doc.build(elems)
    buf.seek(0)
    return buf.read()
