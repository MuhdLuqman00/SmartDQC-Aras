"""
SmartDQC API — FastAPI application
All endpoints. Business logic lives in backend/eda/, backend/cleaning/, and backend/export/.
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
import uuid
from contextlib import asynccontextmanager

logger = logging.getLogger(__name__)

import pandas as pd
import numpy as np
from fastapi import (
    Depends,
    FastAPI,
    File,
    Form,
    Header,
    HTTPException,
    Query,
    UploadFile,
)
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, Response
from pydantic import BaseModel, Field
from typing import List, Optional

from .config import (
    STANDARD_SCHEMA,
    auto_suggest_mapping,
    detect_source_type,
    normalize_schema_type,
)
from .ai.schema_mapper import ai_suggest_mapping, _needs_ai_assist
from .eda.runner import run_eda, run_eda_auto, json_safe
from .eda.kkm_quality_rules import analyze_kkm_quality
from .eda.charts import build_chart_blocks
from .export.tableau import (
    build_aggregated_table,
    to_excel as tbl_excel,
    to_csv as tbl_csv,
)
from .export.cleaned import (
    to_excel as cln_excel,
    to_csv as cln_csv,
    to_excel_typed as cln_excel_typed,
)
from .auth import (
    hash_password,
    verify_password,
    create_access_token,
)
from .db.init_db import init_db, get_db
from .db.models import Dataset, Session as DBSession, AnalysisResult, User
from .ai.narrative import generate_narrative
from .ai.nlq import answer_query
from .ai.ollama_client import OllamaError
from .ml.corrections import flag_anomalies
from .ml.risk_score import compute_risk_scores
from .export.report import build_pptx_bytes, build_pdf_bytes
from .eda.kpi import (
    compute_kpi_dashboard,
    compute_trajectory_narratives,
    compute_district_period_snapshots,
    official_targets,
)

from datetime import datetime
from sqlalchemy.orm import Session as SASession


@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    _seed_admin()
    # Warm the Ollama model in the background so the first AI-insight click
    # (often after the box has sat idle) doesn't pay a cold model-load and 500.
    import asyncio
    from .ai.ollama_client import warmup as _ollama_warmup

    asyncio.get_running_loop().run_in_executor(None, _ollama_warmup)
    yield


def _seed_admin():
    from .db.init_db import SessionLocal

    db = SessionLocal()
    try:
        if not db.query(User).filter_by(username="admin").first():
            db.add(
                User(
                    username="admin",
                    password_hash=hash_password(
                        os.environ.get(
                            "ADMIN_SEED_PASSWORD", "ADMIN_SEED_PASSWORD_PLACEHOLDER"
                        )
                    ),
                    role="admin",
                )
            )
            db.commit()
    finally:
        db.close()


app = FastAPI(title="SmartDQC API", version="1.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── FILE READER ──────────────────────────────────────────────────────────────


def read_file(content: bytes, filename: str, sheet: str | None = None):
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        buf = io.BytesIO(content)
        if len(content) > 50_000_000:
            chunks = pd.read_csv(buf, dtype=str, chunksize=100_000)
            return pd.concat(chunks, ignore_index=True), None
        return pd.read_csv(buf, dtype=str), None
    elif fn.endswith((".xlsx", ".xls")):
        xl = pd.ExcelFile(io.BytesIO(content))
        sheets = xl.sheet_names
        target = sheet if sheet and sheet in sheets else sheets[0]
        return pd.read_excel(io.BytesIO(content), sheet_name=target, dtype=str), sheets
    raise ValueError("Hanya CSV, XLSX, XLS disokong.")


def _csv_stream(df: pd.DataFrame, chunksize: int = 10_000):
    yield df.iloc[0:0].to_csv(index=False)
    for start in range(0, len(df), chunksize):
        yield df.iloc[start : start + chunksize].to_csv(index=False, header=False)


# ─── HEALTH CHECK ─────────────────────────────────────────────────────────────


@app.get("/")
def root():
    return {"message": "SmartDQC API", "version": "1.0", "status": "ok"}


@app.get("/health")
def health():
    return {"status": "ok"}


# ─── AUTH ENDPOINTS ───────────────────────────────────────────────────────────


@app.post("/auth/login")
def login(form: OAuth2PasswordRequestForm = Depends(), db=Depends(get_db)):
    user = db.query(User).filter_by(username=form.username, is_active=True).first()
    if not user or not verify_password(form.password, user.password_hash):
        _log_audit("login_failed", detail=f"username={form.username}")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_access_token({"sub": user.username, "role": user.role})
    _log_audit("login", user_id=user.id, detail=f"role={user.role}")
    return {"access_token": token, "token_type": "bearer", "role": user.role}


def _identity(x_user: str | None = Header(None, alias="X-User")) -> str | None:
    """Anonymous named-identity: the name the user typed, sent as the X-User
    header. There is no password/token — access control is the deployment's
    network perimeter (office LAN + Tailscale/ZeroTier). Used to scope the
    dataset library / sessions per person and to attribute audit entries.

    Normalized to lowercase so the same person typing "Ben" on one device and
    "ben" on another resolves to the same owner — cross-device history is the
    whole point, and case-sensitivity would silently split it. The frontend
    keeps the original casing in localStorage for display. Returns None when
    the header is absent or blank."""
    v = (x_user or "").strip().lower()
    return v or None


def _current_user(x_user: str | None = Header(None, alias="X-User")):
    """Back-compat shim after the password login was removed. Returns a synthetic
    identity built from the X-User name; no token required. Everyone is treated
    as admin (see require_admin) because access is gated at the network layer."""
    from types import SimpleNamespace

    name = (x_user or "").strip() or "anonymous"
    return SimpleNamespace(username=name, role="admin", id=None)


def require_admin(user=Depends(_current_user)):
    """Formerly an admin-only gate. With the password login removed and access
    controlled at the network layer, every identified user is allowed. Kept as a
    pass-through so endpoints depending on it keep working without a token."""
    return user


@app.get("/auth/me")
def me(user=Depends(_current_user)):
    return {"username": user.username, "role": user.role}


@app.post("/auth/logout")
def logout():
    return {"detail": "logged out"}


@app.get("/schema")
def get_schema():
    return {"schema": STANDARD_SCHEMA}


@app.get("/data-dictionary")
def data_dictionary(fmt: str = Query("json", pattern="^(json|excel|pdf)$")):
    derived = {
        "id_cleaned": {
            "type": "identifier",
            "derived": True,
            "description": "IC/MyKid selepas dibersih (12 digit)",
        },
        "id_type": {
            "type": "categorical",
            "derived": True,
            "description": "Jenis ID: valid_ic / short_id / system_id / missing / invalid_ic",
        },
        "is_valid_ic": {
            "type": "boolean",
            "derived": True,
            "description": "True = IC 12 digit sah",
        },
        "age_months_computed": {
            "type": "numerical",
            "derived": True,
            "description": "Umur dalam bulan (dikira dari tarikh_lahir dan tarikh_ukur)",
        },
        "age_group_computed": {
            "type": "categorical",
            "derived": True,
            "description": "Kumpulan umur standard WHO",
        },
        "bulan_ukur": {
            "type": "categorical",
            "derived": True,
            "description": "Bulan pengukuran (1–12)",
        },
        "suku_tahun": {
            "type": "categorical",
            "derived": True,
            "description": "Suku tahun pengukuran (S1–S4)",
        },
        "kawasan_bahagian": {
            "type": "categorical",
            "derived": True,
            "description": "Kawasan (Sabah) / Bahagian (Sarawak)",
        },
        "flag_bawah_2": {
            "type": "boolean",
            "derived": True,
            "description": "True = bawah 2 tahun (< 24 bulan)",
        },
        "flag_bawah_5": {
            "type": "boolean",
            "derived": True,
            "description": "True = bawah 5 tahun (< 60 bulan)",
        },
        "is_missing_critical": {
            "type": "boolean",
            "derived": True,
            "description": "True = berat/tinggi/BMI hilang",
        },
        "is_missing_age": {
            "type": "boolean",
            "derived": True,
            "description": "True = umur tidak dapat dikira",
        },
        "flag_bmi_mismatch": {
            "type": "boolean",
            "derived": True,
            "description": "True = BMI tidak konsisten dengan berat/tinggi (threshold ±1.0)",
        },
        "flag_date_invalid": {
            "type": "boolean",
            "derived": True,
            "description": "True = tarikh_ukur < tarikh_lahir (tidak logik)",
        },
        "flag_date_future": {
            "type": "boolean",
            "derived": True,
            "description": "True = tarikh_ukur dalam masa hadapan",
        },
        "status_bmi_grouped": {
            "type": "categorical",
            "derived": True,
            "description": "BMI grouped: susut/normal/kurang/obes_berlebihan",
        },
        # WHO z-scores
        "waz": {
            "type": "numerical",
            "derived": True,
            "description": "WHO 2006 Weight-for-Age Z-score",
        },
        "haz": {
            "type": "numerical",
            "derived": True,
            "description": "WHO 2006 Height-for-Age Z-score",
        },
        "baz": {
            "type": "numerical",
            "derived": True,
            "description": "WHO 2006 BMI-for-Age Z-score",
        },
        "waz_class": {
            "type": "categorical",
            "derived": True,
            "description": "Klasifikasi WAZ (WHO 2006)",
        },
        "haz_class": {
            "type": "categorical",
            "derived": True,
            "description": "Klasifikasi HAZ (WHO 2006)",
        },
        "baz_class": {
            "type": "categorical",
            "derived": True,
            "description": "Klasifikasi BAZ (WHO 2006)",
        },
        # Indicator flags (z-score based)
        "ind_kurang_berat_zscore": {
            "type": "boolean",
            "derived": True,
            "description": "Kurang berat badan (WAZ < -2, WHO 2006)",
        },
        "ind_bantut_zscore": {
            "type": "boolean",
            "derived": True,
            "description": "Bantut (HAZ < -2, WHO 2006)",
        },
        "ind_susut_zscore": {
            "type": "boolean",
            "derived": True,
            "description": "Susut (BAZ < -2, WHO 2006)",
        },
        "ind_obes_zscore": {
            "type": "boolean",
            "derived": True,
            "description": "Berlebihan berat badan / obes (BAZ > +2, WHO 2006)",
        },
        # Mismatch flags
        "flag_status_berat_vs_zscore": {
            "type": "boolean",
            "derived": True,
            "description": "Label sumber status_berat tidak sepadan dengan WAZ z-skor",
        },
        "flag_status_tinggi_vs_zscore": {
            "type": "boolean",
            "derived": True,
            "description": "Label sumber status_tinggi tidak sepadan dengan HAZ z-skor",
        },
        "flag_status_bmi_vs_zscore": {
            "type": "boolean",
            "derived": True,
            "description": "Label sumber status_bmi tidak sepadan dengan BAZ z-skor",
        },
    }
    if fmt == "excel":
        from .export.data_dictionary import to_excel as _dd_excel

        data = _dd_excel(STANDARD_SCHEMA, derived)
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="SmartDQC_Kamus_Data.xlsx"'
            },
        )
    if fmt == "pdf":
        from .export.data_dictionary import to_pdf as _dd_pdf

        data = _dd_pdf(STANDARD_SCHEMA, derived)
        return StreamingResponse(
            iter([data]),
            media_type="application/pdf",
            headers={
                "Content-Disposition": 'attachment; filename="SmartDQC_Kamus_Data.pdf"'
            },
        )
    return {"source_fields": STANDARD_SCHEMA, "derived_fields": derived}


# ─── UPLOAD / PREVIEW ─────────────────────────────────────────────────────────


@app.post("/upload/preview")
async def upload_preview(
    file: UploadFile = File(...),
    source_type: str = Form("auto"),
    sheet: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=5, le=100),
):
    content = await file.read()
    filename = file.filename or ""
    try:
        df, sheets = read_file(content, filename, sheet)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"Fail tidak dapat dibaca: {e}")

    # Detect source type and generate mapping
    detected_source_type = (
        detect_source_type(df.columns.tolist())
        if source_type == "auto"
        else source_type
    )

    auto_map = auto_suggest_mapping(df.columns.tolist(), detected_source_type)

    if _needs_ai_assist(auto_map):
        sample_values = {
            col: df[col].dropna().head(3).astype(str).tolist()
            for col in df.columns[:30]
        }
        ai_map = ai_suggest_mapping(
            df.columns.tolist(), sample_values, source_type=detected_source_type
        )
        for field, val in ai_map.items():
            if auto_map.get(field) is None and val is not None:
                auto_map[field] = val

    # Cache the uploaded DataFrame for subsequent cleaning
    cache_id = _cache_cleaned(
        df, {"filename": filename, "source_type": detected_source_type}
    )

    # Calculate unmapped columns (columns not in auto_map keys)
    unmapped = [col for col in df.columns if col not in auto_map]

    total_rows = len(df)
    start = (page - 1) * page_size
    page_data = df.iloc[start : start + page_size].copy()
    page_data = page_data.replace(r"^\s*$", np.nan, regex=True)
    page_data = page_data.where(page_data.notna(), None)

    # Build response matching frontend expectations
    mapping_with_confidence: dict = {}
    for col, standard in auto_map.items():
        # Calculate confidence based on match quality
        confidence = 0.95 if standard else 0.0
        mapping_with_confidence[col] = {"standard": standard, "confidence": confidence}

    return JSONResponse(
        content=json_safe(
            {
                "cache_id": cache_id,
                "filename": filename,
                "source_type": detected_source_type,
                "rows": total_rows,
                "columns": df.columns.tolist(),
                "sample": page_data.to_dict(orient="records"),
                "auto_mapping": mapping_with_confidence,
                "unmapped_columns": unmapped,
                "sheets": sheets or [],
                "active_sheet": sheet or (sheets[0] if sheets else None),
                "page": page,
                "page_size": page_size,
                "total_pages": max(1, (total_rows + page_size - 1) // page_size),
            }
        )
    )


# ─── MAPPING VALIDATION ───────────────────────────────────────────────────────


class MappingBody(BaseModel):
    """JSON body sent by the v2 frontend: { "raw_column": "standard_field", … }."""

    mapping: dict[str, str] = {}
    # E2: optional user-chosen dataset name (defaults to the filename when blank).
    dataset_name: Optional[str] = None
    # B3: rule codes the user kept enabled. None ⇒ all rules (locked always run).
    enabled_rules: Optional[List[str]] = None


def _resolve_cached_df(cache_id: Optional[str]) -> pd.DataFrame:
    """Resolve a DataFrame from the in-memory upload cache by cache_id.

    The v2 frontend uploads once via /upload/preview (which caches the df and
    returns a cache_id), then references that id on every subsequent wizard
    step instead of re-uploading the file.
    """
    if not cache_id:
        raise HTTPException(400, "cache_id is required")
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(
            404, f"Cache ID '{cache_id}' not found — please re-upload the file."
        )
    return entry["df"].copy()


# Stat keys that are survivor counts / transformations, not data-quality issues.
_NON_ISSUE_STAT_KEYS = {"final_count", "raw_count", "valid_records", "total_dropped"}


def _summarise_cleaning(stats: dict, rows_before: int, rows_after: int) -> dict:
    """Derive a quality score, applied-rule list, and top issues from a
    cleaner's stats dict (per-rule removed/flagged row counts).

    Generic across kpm/myvass/ncdc/kkm cleaners: any positive-integer stat
    that isn't a survivor/transformation count is treated as an issue+rule.
    """
    score = round(rows_after / rows_before * 100, 1) if rows_before else 0.0
    score = max(0.0, min(100.0, score))

    issues: list[dict] = []
    for key, val in stats.items():
        if isinstance(val, bool) or not isinstance(val, int) or val <= 0:
            continue
        if key in _NON_ISSUE_STAT_KEYS or key.startswith(
            ("standardized_", "ind_", "gender_")
        ):
            continue
        label = key.replace("_", " ").strip().capitalize()
        pct = (val / rows_before * 100) if rows_before else 0
        severity = "critical" if pct >= 10 else "warning" if pct >= 1 else "info"
        # `code` is the raw stat key (e.g. "dropped_duplicate_ic") so the
        # frontend can localise via a bilingual catalog. `description` stays
        # the English label for back-compat (exports, cached data, fallback).
        issues.append(
            {"code": key, "description": label, "severity": severity, "count": val}
        )

    issues.sort(key=lambda i: i["count"], reverse=True)
    return {
        "quality_score": score,
        # `rules_applied` keeps its original string[] shape for back-compat.
        # `rules` is the additive, localisable parallel (code + English label).
        "rules_applied": [i["description"] for i in issues],
        "rules": [{"code": i["code"], "description": i["description"]} for i in issues],
        "top_issues": issues[:5],
    }


@app.post("/mapping/validate")
async def validate_mapping(
    cache_id: Optional[str] = Query(None),
    body: Optional[MappingBody] = None,
):
    df = _resolve_cached_df(cache_id)

    # Frontend sends { raw_column: standard_field }; the validation logic
    # below expects { standard_field: raw_column }, so invert it.
    raw_to_std = body.mapping if body else {}
    mapping_dict = {std: raw for raw, std in raw_to_std.items() if std}

    available_cols = set(df.columns.tolist())
    errors, warnings, ok = [], [], []

    for std_field, raw_col in mapping_dict.items():
        if not raw_col:
            continue
        if raw_col not in available_cols:
            errors.append(
                {
                    "field": std_field,
                    "mapped_to": raw_col,
                    "issue": f"Kolum '{raw_col}' tidak wujud dalam fail",
                }
            )
        else:
            schema_type = STANDARD_SCHEMA.get(std_field, {}).get("type", "unknown")
            col_dtype = str(df[raw_col].dtype)
            ok.append({"field": std_field, "mapped_to": raw_col, "dtype": col_dtype})
            if schema_type == "numerical" and col_dtype == "object":
                pct_numeric = pd.to_numeric(df[raw_col], errors="coerce").notna().mean()
                if pct_numeric < 0.5:
                    warnings.append(
                        {
                            "field": std_field,
                            "mapped_to": raw_col,
                            "issue": f"Kolum kelihatan bukan numerik ({round(pct_numeric * 100)}% boleh ditukar)",
                        }
                    )

    from .config import CORE_FIELDS

    critical_unmapped = [
        f for f in CORE_FIELDS if f not in mapping_dict or not mapping_dict.get(f)
    ]
    if critical_unmapped:
        warnings.append(
            {
                "field": "critical_fields",
                "mapped_to": None,
                "issue": f"Medan kritikal tidak dipetakan: {critical_unmapped}",
            }
        )

    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "ok": ok,
        "total_mapped": len([v for v in mapping_dict.values() if v]),
        "available_columns": df.columns.tolist(),
    }


# ─── RUN EDA ──────────────────────────────────────────────────────────────────


@app.post("/eda/run")
async def run_eda_endpoint(
    file: UploadFile = File(...),
    mapping: str = "",
    source_type: str = "myvass",
    sheet: Optional[str] = None,
    bmi_threshold: float = Query(1.0, ge=0.1, le=10.0),
    owner: str | None = Depends(_identity),
    db=Depends(get_db),
):
    content = await file.read()
    filename = file.filename or ""
    try:
        df, _ = read_file(content, filename, sheet)
    except Exception as e:
        raise HTTPException(400, str(e))

    try:
        mapping_dict = json.loads(mapping) if mapping else {}
    except Exception:
        mapping_dict = {}

    report = run_eda(df, mapping_dict, source_type, bmi_threshold=bmi_threshold)

    # Cache cleaned DF for downstream endpoints
    cleaned_df_data = report.get("_cleaned_data", [])
    import pandas as _pd

    cleaned_df = _pd.DataFrame(cleaned_df_data) if cleaned_df_data else df
    cache_id = _cache_cleaned(cleaned_df, report)

    persisted = True
    persist_error: Optional[str] = None
    try:
        _persist_session(
            cache_id=cache_id,
            filename=filename,
            source_type=source_type,
            row_count=len(cleaned_df),
            result=report,
            db=db,
            owner=owner,
        )
    except Exception as exc:  # best-effort — never fail the EDA run on a DB error
        persisted = False
        persist_error = f"{type(exc).__name__}: {exc}"
        logger.warning(
            "Session persistence failed on /eda/run for cache_id=%s (%s); "
            "results will not be saved to the dashboard until the database "
            "is reachable.",
            cache_id,
            persist_error,
        )

    _log_audit(action="eda.run", detail=f"cache_id={cache_id}", actor=owner)

    # Strip private / large keys from public response
    for key in ["_cleaned_data", "_cleaned_columns", "_aggregated_full"]:
        report.pop(key, None)

    report["cache_id"] = cache_id
    report["persisted"] = persisted
    report["persist_error"] = persist_error
    return JSONResponse(content=json_safe(report))


# ─── CLEANED DATA PREVIEW (paginated) ────────────────────────────────────────


@app.post("/cleaned/preview")
async def cleaned_preview(
    file: UploadFile = File(...),
    mapping: str = "",
    source_type: str = "myvass",
    sheet: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=5, le=200),
):
    content = await file.read()
    filename = file.filename or ""
    try:
        df, _ = read_file(content, filename, sheet)
    except Exception as e:
        raise HTTPException(400, str(e))

    mapping_dict = json.loads(mapping) if mapping else {}
    report = run_eda(df, mapping_dict, source_type)
    all_records = report.get("_cleaned_data", [])
    cols = report.get("_cleaned_columns", [])
    total_rows = len(all_records)
    start = (page - 1) * page_size

    return JSONResponse(
        content=json_safe(
            {
                "total_rows": total_rows,
                "total_pages": max(1, (total_rows + page_size - 1) // page_size),
                "page": page,
                "page_size": page_size,
                "columns": cols,
                "records": all_records[start : start + page_size],
            }
        )
    )


# ─── DOWNLOAD CLEANED DATA ────────────────────────────────────────────────────


@app.post("/download/cleaned")
async def download_cleaned(
    file: UploadFile = File(...),
    mapping: str = "",
    source_type: str = "myvass",
    sheet: Optional[str] = None,
    fmt: str = Query("csv", pattern="^(csv|xlsx)$"),
):
    content = await file.read()
    filename = file.filename or ""
    try:
        df, _ = read_file(content, filename, sheet)
    except Exception as e:
        raise HTTPException(400, str(e))

    mapping_dict = json.loads(mapping) if mapping else {}
    report = run_eda(df, mapping_dict, source_type)
    cleaned_records = report.get("_cleaned_data", [])
    cleaned_cols = report.get("_cleaned_columns", [])
    if not cleaned_records:
        raise HTTPException(500, "Tiada data dibersih.")

    base = filename.rsplit(".", 1)[0]
    if fmt == "csv":
        data = cln_csv(cleaned_records, cleaned_cols)
        return StreamingResponse(
            iter([data]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="cleaned_{base}.csv"'
            },
        )
    else:
        data = cln_excel(cleaned_records, cleaned_cols, base)
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="cleaned_{base}.xlsx"'
            },
        )


# ─── EXPORT AGGREGATED (TABLEAU) — NO ROW CAP ────────────────────────────────


@app.post("/export/aggregated")
async def export_aggregated(
    file: UploadFile = File(...),
    mapping: str = "",
    source_type: str = "myvass",
    sheet: Optional[str] = None,
    fmt: str = Query("csv", pattern="^(csv|xlsx)$"),
):
    """
    Tableau-ready flat aggregated table — every geo × age_group × indicator row.
    NO row cap. Includes sumber_indikator column (zscore | source_label).
    """
    content = await file.read()
    filename = file.filename or ""
    try:
        df, _ = read_file(content, filename, sheet)
    except Exception as e:
        raise HTTPException(400, str(e))

    mapping_dict = json.loads(mapping) if mapping else {}
    report = run_eda(df, mapping_dict, source_type)

    # Use the full (uncapped) aggregated table from the private key
    agg_rows = report.get("_aggregated_full", build_aggregated_table(report))
    if not agg_rows:
        raise HTTPException(
            422, "Tiada data indikator — pastikan status / berat / tinggi dipetakan."
        )

    base = filename.rsplit(".", 1)[0]
    if fmt == "xlsx":
        data = tbl_excel(agg_rows, base)
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="tableau_{base}.xlsx"'
            },
        )
    else:
        data = tbl_csv(agg_rows)
        return StreamingResponse(
            iter([data]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="tableau_{base}.csv"'
            },
        )


@app.get("/export/aggregated-cached/{cache_id}")
async def export_aggregated_cached(
    cache_id: str,
    fmt: str = Query("csv", pattern="^(csv|xlsx)$"),
):
    """Tableau-ready aggregated table for an already-uploaded dataset (cache_id),
    so the v2 wizard can export without re-uploading the file."""
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(
            404, "cache_id not found — run /clean/run first or check the UUID"
        )
    src = (entry.get("stats") or {}).get("source_type") or "myvass"
    report = run_eda_auto(entry["df"], src)
    agg_rows = report.get("_aggregated_full", build_aggregated_table(report))
    if not agg_rows:
        raise HTTPException(
            422, "Tiada data indikator — pastikan status / berat / tinggi dipetakan."
        )

    base = ((entry.get("stats") or {}).get("filename") or "dataset").rsplit(".", 1)[0]
    if fmt == "xlsx":
        data = tbl_excel(agg_rows, base)
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="tableau_{base}.xlsx"'
            },
        )
    data = tbl_csv(agg_rows)
    return StreamingResponse(
        iter([data]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="tableau_{base}.csv"'},
    )


# ─── MYVASS WIDE-TO-LONG PREVIEW ─────────────────────────────────────────────


@app.post("/transform/myvass-wide-to-long")
async def myvass_wide_to_long_endpoint(
    file: UploadFile = File(...),
    sheet: Optional[str] = None,
):
    content = await file.read()
    filename = file.filename or ""
    try:
        df, _ = read_file(content, filename, sheet)
    except Exception as e:
        raise HTTPException(400, str(e))

    from .eda.runner import myvass_wide_to_long

    long_df = myvass_wide_to_long(df)
    return JSONResponse(
        content=json_safe(
            {
                "original_rows": len(df),
                "long_rows": len(long_df),
                "columns": long_df.columns.tolist(),
                "preview": long_df.head(10)
                .replace({np.nan: None})
                .to_dict(orient="records"),
            }
        )
    )


# ═══════════════════════════════════════════════════════════════════════════════
# MYVASS MULTI-FILE MERGE ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════


def _merge_myvass_files(
    file_contents: list[tuple[str, bytes]],
    dose_date_col: str = "DOSE_DATE",
    ic_col: str = "IC_NO_PASSPORT",
) -> tuple[pd.DataFrame, dict]:
    """
    Merge multiple MyVASS files:
    1. Validate all headers match
    2. Concatenate
    3. Deduplicate by IC — keep row with latest DOSE_DATE
    4. Remove rows with any null cell
    Returns (merged_df, merge_stats).
    """
    if not file_contents:
        raise ValueError("No files provided.")

    dfs = []
    ref_cols = None
    ref_name = None

    for filename, content in file_contents:
        fn = filename.lower()
        if fn.endswith(".csv"):
            df = pd.read_csv(io.BytesIO(content), dtype=str)
        elif fn.endswith((".xlsx", ".xls")):
            df = pd.read_excel(io.BytesIO(content), dtype=str)
        else:
            raise ValueError(f"Unsupported format: {filename}")

        if ref_cols is None:
            ref_cols = list(df.columns)
            ref_name = filename
        else:
            if list(df.columns) != ref_cols:
                missing = set(ref_cols) - set(df.columns)
                extra = set(df.columns) - set(ref_cols)
                raise ValueError(
                    f"Header mismatch in '{filename}' vs '{ref_name}'. "
                    f"Missing: {missing or 'none'}. Extra: {extra or 'none'}."
                )
        dfs.append((filename, df))

    merged = pd.concat([df for _, df in dfs], ignore_index=True)
    total_before = len(merged)
    file_counts = {fn: len(df) for fn, df in dfs}

    # Dedup by IC — keep latest DOSE_DATE
    if ic_col not in merged.columns:
        raise ValueError(
            f"Column '{ic_col}' not found. Available: {list(merged.columns)}"
        )

    dup_removed = 0
    if dose_date_col in merged.columns:
        merged["_dd_parsed"] = pd.to_datetime(
            merged[dose_date_col], dayfirst=True, errors="coerce"
        )
        merged = merged.sort_values("_dd_parsed", ascending=False, na_position="last")
        before_dedup = len(merged)
        merged = merged.drop_duplicates(subset=[ic_col], keep="first")
        dup_removed = before_dedup - len(merged)
        merged = merged.drop(columns=["_dd_parsed"]).sort_index().reset_index(drop=True)
    else:
        before_dedup = len(merged)
        merged = merged.drop_duplicates(subset=[ic_col], keep="first")
        dup_removed = before_dedup - len(merged)

    after_dedup = len(merged)

    # Remove rows with any null
    # First replace empty/whitespace strings with NaN for proper null detection
    merged = merged.replace(r"^\s*$", np.nan, regex=True)
    null_rows = int(merged.isna().any(axis=1).sum())
    merged = merged.dropna().reset_index(drop=True)
    after_null_removal = len(merged)

    stats = {
        "files_merged": len(dfs),
        "file_counts": file_counts,
        "total_rows_before": total_before,
        "duplicates_removed": dup_removed,
        "after_dedup": after_dedup,
        "null_rows_removed": null_rows,
        "final_rows": after_null_removal,
        "columns": list(merged.columns),
        "ic_column": ic_col,
        "dose_date_column": dose_date_col if dose_date_col in ref_cols else None,
    }
    return merged, stats


@app.post("/upload/merge-preview")
async def merge_preview(
    files: List[UploadFile] = File(...),
    ic_col: str = Query("IC_NO_PASSPORT"),
    dose_date_col: str = Query("DOSE_DATE"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=5, le=100),
):
    """
    Upload multiple MyVASS files, merge + dedup + remove nulls,
    and return a preview of the merged data.
    """
    file_contents = []
    for f in files:
        content = await f.read()
        file_contents.append((f.filename or "unknown", content))

    try:
        merged, stats = _merge_myvass_files(file_contents, dose_date_col, ic_col)
    except ValueError as e:
        raise HTTPException(400, str(e))

    source_type = detect_source_type(merged.columns.tolist())
    auto_map = auto_suggest_mapping(merged.columns.tolist(), source_type)

    if _needs_ai_assist(auto_map):
        sample_values = {
            col: merged[col].dropna().head(3).astype(str).tolist()
            for col in merged.columns[:30]
        }
        ai_map = ai_suggest_mapping(
            merged.columns.tolist(), sample_values, source_type=source_type
        )
        for field, val in ai_map.items():
            if auto_map.get(field) is None and val is not None:
                auto_map[field] = val

    total_rows = len(merged)
    start = (page - 1) * page_size
    page_data = merged.iloc[start : start + page_size].copy()
    page_data = page_data.where(page_data.notna(), None)

    # Save merged data in memory as CSV for subsequent EDA call
    merged_buf = io.BytesIO()
    merged.to_csv(merged_buf, index=False, encoding="utf-8-sig")
    merged_bytes = merged_buf.getvalue()

    _log_audit(action="upload.preview", detail=",".join(f[0] for f in file_contents))
    return JSONResponse(
        content=json_safe(
            {
                "merge_stats": stats,
                "source_type": source_type,
                "total_rows": total_rows,
                "total_columns": len(merged.columns),
                "columns": merged.columns.tolist(),
                "preview": page_data.to_dict(orient="records"),
                "page": page,
                "page_size": page_size,
                "total_pages": max(1, (total_rows + page_size - 1) // page_size),
                "auto_mapping": auto_map,
            }
        )
    )


@app.post("/eda/run-merged")
async def run_eda_merged(
    files: List[UploadFile] = File(...),
    mapping: str = "",
    source_type: str = "myvass",
    ic_col: str = Query("IC_NO_PASSPORT"),
    dose_date_col: str = Query("DOSE_DATE"),
    bmi_threshold: float = Query(1.0, ge=0.1, le=10.0),
):
    """
    Upload multiple MyVASS files → merge + dedup + remove nulls → run full EDA.
    """
    file_contents = []
    for f in files:
        content = await f.read()
        file_contents.append((f.filename or "unknown", content))

    try:
        merged, merge_stats = _merge_myvass_files(file_contents, dose_date_col, ic_col)
    except ValueError as e:
        raise HTTPException(400, str(e))

    try:
        mapping_dict = json.loads(mapping) if mapping else {}
    except Exception:
        mapping_dict = {}

    # Convert merged DataFrame to string-typed (same as read_file does)
    merged = merged.astype(str).replace("nan", np.nan)

    report = run_eda(merged, mapping_dict, source_type, bmi_threshold=bmi_threshold)
    report["merge_stats"] = merge_stats
    report["changes_applied"].insert(
        0,
        f"Merged {merge_stats['files_merged']} files: "
        f"{merge_stats['total_rows_before']} total rows → "
        f"{merge_stats['duplicates_removed']} duplicates removed → "
        f"{merge_stats['null_rows_removed']} null rows removed → "
        f"{merge_stats['final_rows']} final rows",
    )

    for key in ["_cleaned_data", "_cleaned_columns", "_aggregated_full"]:
        report.pop(key, None)

    return JSONResponse(content=json_safe(report))


@app.post("/download/cleaned-merged")
async def download_cleaned_merged(
    files: List[UploadFile] = File(...),
    mapping: str = "",
    source_type: str = "myvass",
    ic_col: str = Query("IC_NO_PASSPORT"),
    dose_date_col: str = Query("DOSE_DATE"),
    fmt: str = Query("csv", pattern="^(csv|xlsx)$"),
):
    """Download cleaned data from merged multi-file upload."""
    file_contents = []
    for f in files:
        content = await f.read()
        file_contents.append((f.filename or "unknown", content))

    try:
        merged, _ = _merge_myvass_files(file_contents, dose_date_col, ic_col)
    except ValueError as e:
        raise HTTPException(400, str(e))

    mapping_dict = json.loads(mapping) if mapping else {}
    merged = merged.astype(str).replace("nan", np.nan)
    report = run_eda(merged, mapping_dict, source_type)
    cleaned_records = report.get("_cleaned_data", [])
    cleaned_cols = report.get("_cleaned_columns", [])
    if not cleaned_records:
        raise HTTPException(500, "Tiada data dibersih.")

    if fmt == "csv":
        data = cln_csv(cleaned_records, cleaned_cols)
        return StreamingResponse(
            iter([data]),
            media_type="text/csv",
            headers={
                "Content-Disposition": 'attachment; filename="cleaned_merged_myvass.csv"'
            },
        )
    else:
        data = cln_excel(cleaned_records, cleaned_cols, "merged_myvass")
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="cleaned_merged_myvass.xlsx"'
            },
        )


# ═══════════════════════════════════════════════════════════════════════════════
# KKM DATA CLEANING TOOL ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

import uuid as _uuid

from .eda.cleaning import (
    clean_data,
    detect_data_type,
    EVALUATED_RULES,
    RULE_REGISTRY,
    LOCKED_RULES,
    rules_for_source,
    REVIEW_RULE_REGISTRY,
    REVIEW_EVALUATED_RULES,
    review_rules_for_source,
    _REVIEW_MANAGED_SENTINEL,
)

# ── Cleaned-DataFrame cache: in-memory hot tier + durable disk tier ──────────
# The v2 frontend references uploaded/cleaned data by cache_id across many
# steps (wizard, Explorer, KPI map, AI). An in-memory dict alone loses every
# cache_id on restart, breaking those pages. Persist each entry to disk so a
# cache_id survives a process restart and is rehydrated on demand.
import os as _os
import pickle as _pickle
from pathlib import Path as _Path

_cleaned_cache: dict[str, dict] = {}  # key -> {"df": DataFrame, "stats": dict}
_CACHE_MAX = 5  # keep at most 5 entries hot in memory
_DISK_CACHE_MAX = 100  # keep at most 100 entries on disk

_CACHE_DIR = _Path(
    _os.environ.get("SMARTDQC_CACHE_DIR")
    or (_Path(__file__).resolve().parent.parent / "data" / "cache")
)
try:
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
except Exception as _exc:  # pragma: no cover - best effort
    logger.warning("Could not create cache dir %s: %s", _CACHE_DIR, _exc)


def _cache_path(key: str) -> "_Path":
    return _CACHE_DIR / f"{key}.pkl"


def _prune_disk_cache() -> None:
    """Keep the disk cache bounded by deleting the oldest files."""
    try:
        files = sorted(_CACHE_DIR.glob("*.pkl"), key=lambda p: p.stat().st_mtime)
        for stale in files[:-_DISK_CACHE_MAX] if len(files) > _DISK_CACHE_MAX else []:
            stale.unlink(missing_ok=True)
    except Exception as exc:  # pragma: no cover - best effort
        logger.warning("Disk cache prune failed: %s", exc)


def _cache_cleaned(df: pd.DataFrame, stats: dict | None = None) -> str:
    """Store cleaned DF + stats in the cache (memory + disk), return a key."""
    key = str(_uuid.uuid4())
    if len(_cleaned_cache) >= _CACHE_MAX:
        _cleaned_cache.pop(next(iter(_cleaned_cache)), None)
    payload = {"df": df, "stats": stats or {}}
    _cleaned_cache[key] = payload
    try:
        with _cache_path(key).open("wb") as fh:
            _pickle.dump(payload, fh, protocol=_pickle.HIGHEST_PROTOCOL)
        _prune_disk_cache()
    except Exception as exc:  # pragma: no cover - best effort
        logger.warning("Cache persist failed for %s: %s", key, exc)
    return key


def _is_valid_cache_key(key: str) -> bool:
    """Cache keys are always uuid4 strings. Reject anything else so a
    user-supplied cache_id can never escape _CACHE_DIR (path traversal →
    pickle.load of an arbitrary file)."""
    try:
        _uuid.UUID(str(key))
        return True
    except (ValueError, AttributeError, TypeError):
        return False


def _cache_get(key: str) -> dict | None:
    """Resolve a cache entry by key: memory first, then rehydrate from disk."""
    if not _is_valid_cache_key(key):
        return None
    if key in _cleaned_cache:
        return _cleaned_cache[key]
    path = _cache_path(key)
    if path.exists():
        try:
            with path.open("rb") as fh:
                payload = _pickle.load(fh)
            if len(_cleaned_cache) >= _CACHE_MAX:
                _cleaned_cache.pop(next(iter(_cleaned_cache)), None)
            _cleaned_cache[key] = payload  # rehydrate hot tier
            return payload
        except Exception as exc:  # pragma: no cover - best effort
            logger.warning("Cache load failed for %s: %s", key, exc)
    return None


def _persist_child_records(
    db, dataset_id: str, source_type: str, recs: list[dict]
) -> int:
    """Persist child records to the child_record table. Idempotent replace."""
    from backend.db.models import ChildRecord

    # Delete existing records for this dataset (idempotent replace)
    db.query(ChildRecord).filter(ChildRecord.dataset_id == dataset_id).delete()
    # Bulk insert
    rows = [
        ChildRecord(
            dataset_id=dataset_id,
            source_type=normalize_schema_type(source_type or "general"),
            ic_norm=_normalise_ic(r.get("ic", "")),
            name=r.get("name"),
            dob=r.get("dob"),
            gender=r.get("gender"),
            state=r.get("state"),
            district=r.get("district"),
            measure_date=r.get("measure_date"),
            weight_kg=r.get("weight_kg"),
            height_cm=r.get("height_cm"),
            bmi=r.get("bmi"),
            waz=r.get("waz"),
            haz=r.get("haz"),
            baz=r.get("baz"),
        )
        for r in recs
    ]
    db.add_all(rows)
    db.commit()
    return len(rows)


def _records_from_store(
    db, dataset_ids: list[str] | None = None
) -> tuple[list[dict], dict]:
    """P2-3: Read child records from the durable child_record table.

    Returns (records, dataset_created_at_by_id) matching _records_from_cached shape.
    """
    from backend.db.models import ChildRecord, Dataset

    query = db.query(ChildRecord)
    if dataset_ids:
        query = query.filter(ChildRecord.dataset_id.in_(dataset_ids))
    rows = query.all()

    # Build dataset lookup for created_at
    ds_ids = {r.dataset_id for r in rows}
    ds_lookup = {
        ds.id: ds.created_at
        for ds in db.query(Dataset).filter(Dataset.id.in_(ds_ids)).all()
    }

    out = []
    for r in rows:
        rec = {
            "ic": r.ic_norm,
            "source_type": r.source_type,
            "dataset_id": r.dataset_id,
            "name": r.name,
            "dob": r.dob,
            "gender": r.gender,
            "state": r.state,
            "district": r.district,
            "measure_date": r.measure_date,
            "weight_kg": r.weight_kg,
            "height_cm": r.height_cm,
            "bmi": r.bmi,
            "waz": r.waz,
            "haz": r.haz,
            "baz": r.baz,
        }
        if r.dataset_id in ds_lookup:
            rec["_dataset_created_at"] = ds_lookup[r.dataset_id]
        out.append(rec)

    return out, {
        ds_id: dt.isoformat() if dt else None for ds_id, dt in ds_lookup.items()
    }


def _reconciliation_summary(groups: list[dict]) -> dict:
    """P2-5: Compute reconciliation summary from linkage groups."""
    total_records = sum(len(g["sources"]) for g in groups)
    unique_children = len(groups)
    duplicate_records = total_records - unique_children
    duplication_rate = duplicate_records / total_records if total_records > 0 else 0.0

    multi_source_children = sum(
        1 for g in groups if len({s["source_type"] for s in g["sources"]}) >= 2
    )

    # Source overlap: group counts by sorted tuple of source types
    source_overlap: dict[str, int] = {}
    for g in groups:
        sources = tuple(sorted({s["source_type"] for s in g["sources"]}))
        key = "+".join(sources)
        source_overlap[key] = source_overlap.get(key, 0) + 1

    # Conflicts by severity
    conflicts_by_severity = {"hard": 0, "soft": 0, "strong": 0}
    for g in groups:
        severities = {c["severity"] for c in g.get("conflicts", [])}
        for sev in severities:
            if sev in conflicts_by_severity:
                conflicts_by_severity[sev] += 1

    return {
        "total_records": total_records,
        "unique_children": unique_children,
        "duplicate_records": duplicate_records,
        "duplication_rate": round(duplication_rate, 4),
        "multi_source_children": multi_source_children,
        "source_overlap": source_overlap,
        "conflicts_by_severity": conflicts_by_severity,
    }


def _persist_linkage_run(
    db,
    groups: list[dict],
    params: dict,
    dataset_ids: list[str],
    created_by: int | None = None,
) -> int:
    """P2-4: Persist linkage run for audit."""
    from backend.db.models import LinkageRun, LinkageMember

    run = LinkageRun(
        params_json=params,
        dataset_ids=dataset_ids,
        total_groups=len(groups),
        linked_groups=sum(1 for g in groups if len(g["sources"]) > 1),
        created_by=created_by,
    )
    db.add(run)
    db.flush()

    members = []
    for gi, g in enumerate(groups):
        for src in g["sources"]:
            members.append(
                LinkageMember(
                    run_id=run.id,
                    group_index=gi,
                    ic_norm=_normalise_ic(src.get("ic", "")),
                    source_type=src["source_type"],
                    dataset_id=src["dataset_id"],
                    name=src.get("name"),
                    dob=src.get("dob"),
                    confidence=g.get("confidence"),
                    match_reasons=g.get("match_reasons"),
                )
            )

    if members:
        db.add_all(members)
    db.commit()
    return run.id


# Clinical bounds mirrored from backend/cleaning/kkm.py — keep in sync.
_FLAG_BERAT_LOW, _FLAG_BERAT_HIGH = 12.0, 50.0  # BERAT_MIN / BERAT_MAX
_FLAG_TINGGI_LOW, _FLAG_TINGGI_HIGH = 100.0, 160.0  # TINGGI_MIN / TINGGI_MAX
_FLAG_BMI_LOW, _FLAG_BMI_HIGH = 13.5, 18.5  # BMI_UNDERWEIGHT / BMI_OBESE


def _compute_row_flags(df: "pd.DataFrame") -> "pd.Series":
    """Return a boolean Series: True if the row has a known quality issue.

    Priority: use Data_Quality_Flag column if present (KKM cleaned output).
    Fallback: check clinical bounds on recognised measurement columns so the
    toggle works for non-KKM source types too.
    """
    if "Data_Quality_Flag" in df.columns:
        return df["Data_Quality_Flag"] != "Valid"

    flagged = pd.Series(False, index=df.index)
    if "review_reason" in df.columns:
        flagged |= df["review_reason"].fillna("").astype(str).str.strip().ne("")
    for col in df.columns:
        c = col.lower()
        if "berat" in c and "kg" in c:
            vals = pd.to_numeric(df[col], errors="coerce")
            flagged |= (
                vals.isna() | (vals < _FLAG_BERAT_LOW) | (vals > _FLAG_BERAT_HIGH)
            )
        elif "tinggi" in c and "cm" in c:
            vals = pd.to_numeric(df[col], errors="coerce")
            flagged |= (
                vals.isna() | (vals < _FLAG_TINGGI_LOW) | (vals > _FLAG_TINGGI_HIGH)
            )
        elif "bmi" in c:
            vals = pd.to_numeric(df[col], errors="coerce")
            flagged |= vals.isna() | (vals < _FLAG_BMI_LOW) | (vals > _FLAG_BMI_HIGH)
    return flagged


# ── Flag-then-filter view projections ────────────────────────────────────────
# Cleaners now return the FULL frame plus two bookkeeping columns:
#   analyzable     (bool) — row passed every quality rule
#   exclude_reason (str)  — semicolon-joined rule codes that excluded the row
# These are internal bookkeeping, not user-facing data. The cache holds the full
# frame (source of truth for KPI, reports, the "full" download), and each read
# boundary projects to the view it needs.
_FLAG_COLS = ("analyzable", "exclude_reason")


def _analysis_view(df: "pd.DataFrame") -> "pd.DataFrame":
    """Analysis projection: analyzable rows only, both flag columns dropped.

    Reproduces the pre-flag-then-filter (drop-based) export exactly — the
    cleaners used to physically drop these rows, so this is what downloads and
    analytics consumed before the refactor. Frames without the flag columns
    (old callers / already-clean frames) pass through unchanged.
    """
    if "analyzable" in df.columns:
        df = df[df["analyzable"]]
    drop = [c for c in _FLAG_COLS if c in df.columns]
    return df.drop(columns=drop) if drop else df


def _download_view(df: "pd.DataFrame", view: str) -> "pd.DataFrame":
    """Project a cached frame for download. ``full`` keeps every row + both
    flag columns; ``analysis`` (default) returns clean rows only."""
    return df if view == "full" else _analysis_view(df)


def _view_label(view: str) -> str:
    """Filename infix that distinguishes the two download views so the files
    don't collide on disk (both used to be named ``..._Cleaned_...``)."""
    return "Full_Flagged" if view == "full" else "Cleaned"


def _view_sheet(view: str) -> str:
    """Excel sheet name matching the download view."""
    return "Full + Flagged" if view == "full" else "Cleaned Data"


def _explorer_view(df: "pd.DataFrame") -> "pd.DataFrame":
    """Explorer projection: keep ALL rows (Trust→Correct→Export, and because
    row-edit ids are positional into the cached frame — dropping rows would
    desync them). Surface ``exclude_reason`` (the per-row 'why'), hide the
    redundant ``analyzable`` bool."""
    return df.drop(columns=["analyzable"]) if "analyzable" in df.columns else df


def _cache_set_narrative(key: str, narrative: dict) -> None:
    """Persist the AI narrative alongside the cleaned-data cache entry.

    Keyed by cache_id and colocated with {df,stats} so the existing
    dataset-delete eviction already covers it. Regeneration overwrites
    (latest wins). Re-pickles to disk so the report endpoints — which may
    rehydrate from disk in a fresh process — see it too.
    """
    entry = _cache_get(key)
    if entry is None:
        return
    entry["narrative"] = narrative
    _cleaned_cache[key] = entry
    try:
        with _cache_path(key).open("wb") as fh:
            _pickle.dump(entry, fh, protocol=_pickle.HIGHEST_PROTOCOL)
    except Exception as exc:  # pragma: no cover - best effort
        logger.warning("Narrative persist failed for %s: %s", key, exc)


def _get_or_build_narrative(key: str, entry: dict) -> dict:
    """Return the cached narrative for this cache_id, or generate + persist
    it once on miss so a report download never silently omits insights.
    If generation fails (e.g. Ollama down) the report still builds — the
    narrative sections are simply omitted rather than 500-ing the download.
    """
    cached = entry.get("narrative")
    if cached:
        return cached
    try:
        source_type = (entry.get("stats") or {}).get("source_type") or "myvass"
        eda_result = run_eda_auto(entry["df"], source_type)
        narrative = generate_narrative(eda_result)
        _cache_set_narrative(key, narrative)
        return narrative
    except Exception as exc:
        logger.warning("Report narrative generation failed for %s: %s", key, exc)
        return {}


def _cache_evict(key: str) -> bool:
    """Remove a cache entry from the in-memory hot tier and the disk tier.
    Returns True if anything was removed."""
    removed = _cleaned_cache.pop(key, None) is not None
    try:
        path = _cache_path(key)
        if path.exists():
            path.unlink(missing_ok=True)
            removed = True
    except Exception as exc:  # pragma: no cover - best effort
        logger.warning("Cache evict failed for %s: %s", key, exc)
    return removed


def _coerce_float(v):
    """Force a true builtin float (or None).

    np.float64 is a *subclass* of Python ``float``, so json_safe()'s
    ``isinstance(obj, float)`` branch returns it UNCHANGED — the value
    stays np.float64. psycopg2 has no adapter for it and falls back to
    NumPy 2.x ``repr`` -> the literal ``np.float64(79.0)`` lands in the
    SQL, and Postgres reads ``np`` as a schema -> InvalidSchemaName.
    A bare float() is the only thing that actually fixes this; do NOT
    "simplify" this back to json_safe().
    """
    return float(v) if v is not None else None


def _persist_session(
    cache_id: str,
    filename: str,
    source_type: str,
    row_count: int,
    result: dict,
    db,
    name: str | None = None,
    owner: str | None = None,
) -> None:
    """Upsert Dataset + Session + AnalysisResult so session data survives server restart."""
    from .db.models import Dataset, Session as _Session, AnalysisResult
    from datetime import datetime as _dt
    from .eda.runner import json_safe

    now = _dt.utcnow()
    has_quality = "quality_score" in result
    quality = _coerce_float(result.get("quality_score"))

    ds = db.query(Dataset).filter_by(id=cache_id).first()
    if ds is None:
        ds = Dataset(
            id=cache_id,
            name=(name or "").strip() or filename,
            filename=filename,
            source_type=source_type,
            row_count=row_count,
            quality_score=quality,
            created_at=now,
            owner=owner,
        )
        db.add(ds)
        db.flush()
    else:
        ds.quality_score = quality if has_quality else ds.quality_score

    sess_id = str(_uuid.uuid4())
    sess = _Session(
        id=sess_id,
        dataset_id=cache_id,
        created_at=now,
        updated_at=now,
    )
    db.add(sess)
    db.flush()

    db.add(
        AnalysisResult(
            id=str(_uuid.uuid4()),
            session_id=sess_id,
            result_type="quality",
            result_json=json_safe(
                {
                    "quality_score": result.get("quality_score"),
                    "issues": result.get("issues", []),
                }
            ),
            created_at=now,
        )
    )
    db.commit()


def _log_audit(
    action: str,
    dataset_id: str | None = None,
    detail: str | None = None,
    user_id: int | None = None,
    actor: str | None = None,
) -> None:
    """Best-effort audit write — logs a warning if it fails but never raises.

    Previously imported from a non-existent `db.session` module so every
    call silently no-op'd — the audit table stayed empty regardless of
    activity. Fixed to use the real SessionLocal from db.init_db.

    `actor` is the anonymous named-identity (X-User) of whoever performed the
    action. With the password login removed there is no users-table user_id to
    attribute by, so the actor name is folded into `detail` (self-asserted —
    accountability, not proof of identity).
    """
    if actor:
        detail = f"actor={actor}" + (f"; {detail}" if detail else "")
    try:
        from .db.init_db import SessionLocal
        from .db.models import AuditLog

        db = SessionLocal()
        try:
            db.add(
                AuditLog(
                    action=action, dataset_id=dataset_id, detail=detail, user_id=user_id
                )
            )
            db.commit()
        finally:
            db.close()
    except Exception as e:
        logger.warning("Audit log write failed for action=%s: %s", action, e)


def _resolve_source(
    file: bytes | None, filename: str | None, cache_id: str | None
) -> "pd.DataFrame":
    """Resolve a join source to a DataFrame from either a raw upload or the cleaned cache."""
    if cache_id is not None:
        entry = _cache_get(cache_id)
        if entry is None:
            raise HTTPException(
                400, f"Cache ID '{cache_id}' not found — re-run cleaning first."
            )
        return entry["df"].copy()
    if file is not None:
        df, _ = read_file(file, filename or "upload.csv")
        return df
    raise HTTPException(
        400, "Provide either a file upload or a cache_id for each side of the join."
    )


def _perform_join(
    df_left: "pd.DataFrame",
    df_right: "pd.DataFrame",
    join_type: str,
    key_cols: list[str] | None,
    dedup: bool,
) -> "tuple[pd.DataFrame, dict]":
    """Execute a join/union and return (result_df, stats_dict)."""
    if join_type == "union":
        result = pd.concat([df_left, df_right], ignore_index=True)
        if dedup:
            before = len(result)
            result = result.drop_duplicates()
            dupes_removed = before - len(result)
        else:
            dupes_removed = 0
        stats = {
            "left_rows": len(df_left),
            "right_rows": len(df_right),
            "result_rows": len(result),
            "duplicates_removed": dupes_removed,
        }
        return result, stats

    # Horizontal join (inner / left / right / outer)
    if not key_cols:
        raise HTTPException(
            400, "key_cols is required for inner/left/right/outer joins."
        )
    missing_left = [c for c in key_cols if c not in df_left.columns]
    missing_right = [c for c in key_cols if c not in df_right.columns]
    if missing_left:
        raise HTTPException(
            400, f"Key column(s) {missing_left} not found in left dataset."
        )
    if missing_right:
        raise HTTPException(
            400, f"Key column(s) {missing_right} not found in right dataset."
        )

    # Match stats based on first key column
    key = key_cols[0]
    left_keys = set(df_left[key].dropna().astype(str))
    right_keys = set(df_right[key].dropna().astype(str))
    stats = {
        "matched_keys": len(left_keys & right_keys),
        "left_only_keys": len(left_keys - right_keys),
        "right_only_keys": len(right_keys - left_keys),
        "result_rows": 0,  # filled after merge
    }

    result = pd.merge(df_left, df_right, on=key_cols, how=join_type)
    stats["result_rows"] = len(result)
    return result, stats


def _profile_columns(df: pd.DataFrame) -> list[dict]:
    """Return column-level profile of a DataFrame."""
    cols = []
    for c in df.columns:
        non_null = int(df[c].notna().sum())
        missing = len(df) - non_null
        missing_pct = round(missing / len(df) * 100, 1) if len(df) > 0 else 0
        unique = int(df[c].nunique())
        # infer type
        s = pd.to_numeric(df[c], errors="coerce")
        dtype = "Numeric" if s.notna().sum() > 0.5 * non_null else "Text"
        cols.append(
            {
                "name": c,
                "non_null": non_null,
                "missing": missing,
                "missing_pct": missing_pct,
                "unique": unique,
                "dtype": dtype,
            }
        )
    return cols


@app.post("/clean/detect-type")
async def detect_type_endpoint(
    file: UploadFile = File(...),
    sheet: Optional[str] = None,
):
    """Detect data type from file columns."""
    content = await file.read()
    filename = file.filename or ""
    try:
        df, sheets = read_file(content, filename, sheet)
    except Exception as e:
        raise HTTPException(400, str(e))

    data_type = detect_data_type(df.columns.tolist(), filename)

    return JSONResponse(
        content=json_safe(
            {
                "filename": filename,
                "detected_type": data_type,
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "columns": df.columns.tolist(),
                "sheets": sheets or [],
            }
        )
    )


# Map KKM business-rule ids → stable, localisable finding codes (issueCatalog.ts).
# Keeps the frontend catalog clean while the backend keeps its BR-xx vocabulary.
_BR_FINDING_CODE = {
    "BR-01": "null_measurement_date",
    "BR-02": "impossible_weight",
    "BR-03": "impossible_height",
    "BR-04": "duplicate_student_id",
    "BR-05": "unknown_gender",
    "BR-06": "unexpected_year_level",
    "BR-07": "dob_in_id",
    "BR-08": "both_measurements_null",
    "BR-09": "suspicious_dates",
}
# KKMQualityChecker severities → frontend severity vocabulary.
_BR_SEVERITY = {
    "CRITICAL": "critical",
    "ERROR": "critical",
    "WARNING": "warning",
    "INFO": "info",
}
_BR_SEVERITY_RANK = {"critical": 0, "warning": 1, "info": 2}


def _actionable_findings(df: pd.DataFrame, limit: int = 6) -> list[dict]:
    """Run the (otherwise unwired) KKM business-rule checker on the RAW frame and
    return a compact, PII-free list of the most actionable findings for B2.1.

    Only aggregate counts/percentages + the rule's own description/fix are
    surfaced — the checker's per-row `affected_rows` (which contain real data)
    are deliberately NOT returned. Defensive: never break quality-check.
    """
    try:
        report = analyze_kkm_quality(df)
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning("Actionable findings skipped (KKM checker failed): %s", exc)
        return []

    findings: list[dict] = []
    for issue in report.get("issues", []):
        rule_id = issue.get("rule_id", "")
        severity = _BR_SEVERITY.get(str(issue.get("severity", "")).upper(), "info")
        findings.append(
            {
                "code": _BR_FINDING_CODE.get(rule_id, rule_id),
                "rule_id": rule_id,
                "field": issue.get("column"),
                "title": issue.get("issue_type"),
                "description": issue.get("description"),  # English detail (expand)
                "fix": issue.get("recommended_fix"),
                "severity": severity,
                "count": int(issue.get("row_count", 0) or 0),
                "pct": issue.get("pct_total", 0),
            }
        )
    findings.sort(key=lambda f: (_BR_SEVERITY_RANK.get(f["severity"], 3), -f["count"]))
    return findings[:limit]


@app.post("/clean/quality-check")
async def quality_check_endpoint(
    cache_id: Optional[str] = Query(None),
    body: Optional[MappingBody] = None,
):
    """Get raw data quality analysis before cleaning (cache_id-based).

    `body` (the proposed mapping) is accepted for contract parity with the
    frontend but not used — quality is computed on the raw cached columns.
    """
    df = _resolve_cached_df(cache_id)

    # Basic quality stats
    quality = {
        "total_rows": len(df),
        "total_columns": len(df.columns),
        "columns": [],
    }

    for col in df.columns:
        col_data = df[col]
        non_null = col_data.notna().sum()
        null_count = len(df) - non_null
        unique_count = col_data.nunique()

        # Check if numeric
        numeric_vals = pd.to_numeric(col_data, errors="coerce")
        is_numeric = numeric_vals.notna().sum() > non_null * 0.8

        col_info = {
            "name": col,
            "non_null": int(non_null),
            "null_count": int(null_count),
            "null_percent": round((null_count / len(df)) * 100, 1)
            if len(df) > 0
            else 0,
            "unique_count": int(unique_count),
            "is_numeric": is_numeric,
        }

        if is_numeric and non_null > 0:
            col_info["min"] = (
                float(numeric_vals.min()) if numeric_vals.notna().any() else None
            )
            col_info["max"] = (
                float(numeric_vals.max()) if numeric_vals.notna().any() else None
            )
            col_info["mean"] = (
                round(float(numeric_vals.mean()), 2)
                if numeric_vals.notna().any()
                else None
            )
        else:
            sample = col_data.dropna().head(5).tolist()
            col_info["sample_values"] = [str(v)[:50] for v in sample]
            # Top categories (B2.1): value_counts of the non-null values, so the
            # user can see WHAT dominates a categorical column at a glance.
            vc = col_data.dropna().value_counts().head(5)
            base = int(non_null) or 1
            col_info["top_values"] = [
                {"value": str(v)[:50], "count": int(c), "pct": round(c / base * 100, 1)}
                for v, c in vc.items()
            ]

        quality["columns"].append(col_info)

    # Overall completeness
    total_cells = len(df) * len(df.columns)
    filled_cells = sum(c["non_null"] for c in quality["columns"])
    quality["overall_completeness"] = (
        round((filled_cells / total_cells) * 100, 1) if total_cells > 0 else 0
    )

    # Frontend Step 3 expects a quality_score + issues list. Use completeness
    # as the pre-clean quality proxy and per-column nulls as the issues.
    col_issues: list[dict] = []
    for c in quality["columns"]:
        nc = int(c.get("null_count", 0) or 0)
        if nc <= 0:
            continue
        np_pct = c.get("null_percent", 0) or 0
        severity = "critical" if np_pct >= 50 else "warning" if np_pct >= 10 else "info"
        col_issues.append(
            {
                # `code` + params let the frontend localise; `description`
                # stays as the English fallback / export string.
                "code": "col_empty",
                "field": c["name"],
                "pct": np_pct,
                "description": f"Column '{c['name']}' is {np_pct}% empty",
                "severity": severity,
                "count": nc,
            }
        )
    col_issues.sort(key=lambda i: i["count"], reverse=True)
    quality["quality_score"] = quality["overall_completeness"]
    quality["issues"] = col_issues[:5]

    # B2.1: prominent, actionable business-rule findings (future dates, dupes,
    # impossible measurements, …) computed on the raw frame — PII-free.
    quality["actionable_findings"] = _actionable_findings(df)

    return JSONResponse(content=json_safe(quality))


@app.post("/clean/run")
async def clean_run_endpoint(
    cache_id: Optional[str] = Query(None),
    data_type: str = Query("myvass"),
    body: Optional[MappingBody] = None,
    owner: str | None = Depends(_identity),
    db=Depends(get_db),
):
    """Run the cleaning process and return cleaned data with statistics.

    The v2 frontend references the previously uploaded data by cache_id.
    `body` (the proposed mapping) is accepted for contract parity.
    """
    df = _resolve_cached_df(cache_id)
    _entry_stats = (_cache_get(cache_id) or {}).get("stats", {}) or {}
    filename = _entry_stats.get("filename", "cached_data")

    # Apply the user-confirmed mapping the wizard already POSTs (was
    # silently ignored): rename raw columns → canonical names BEFORE
    # cleaning, mirroring runner.py's inverse-rename for the EDA path.
    if body and getattr(body, "mapping", None):
        rename = {
            raw: std
            for raw, std in body.mapping.items()
            if std and raw in df.columns and raw != std
        }
        if rename:
            df = df.rename(columns=rename)

    # Drive the cleaner from the detected source_type (cached at upload),
    # not the hardcoded "myvass" default. An explicit non-default data_type
    # from the caller still overrides. Unrecognised/legacy types → general cleaner.
    cached_st = _entry_stats.get("source_type")
    effective_type = normalize_schema_type(
        cached_st if (data_type == "myvass" and cached_st) else data_type
    )

    # D2: build effective rule set = body's drop selection UNION persisted review
    # selection. The upload pipeline sends drop-only codes; review disables must
    # always be merged from Settings so a disabled review rule suppresses its flag
    # even during an upload run. Guard: nothing passed + no review settings = None
    # (all-on default preserved). Best-effort: DB unavailable → all-on.
    _raw_body_rules = (
        set(body.enabled_rules) if (body and body.enabled_rules is not None) else None
    )
    enabled = _effective_enabled_rules(_raw_body_rules, db)
    try:
        cleaned_df, stats = clean_data(df, effective_type, enabled)
    except Exception as e:
        raise HTTPException(500, f"Cleaning error: {str(e)}")

    summary = _summarise_cleaning(
        stats, len(df), stats.get("final_count", len(cleaned_df))
    )

    # Authoritative "Data Quality Score" = the 7-dimension rubric (the same
    # one the AI narrative and report use), NOT the row-survival ratio
    # _summarise_cleaning produces. Scoring it here makes the dashboard,
    # sessions list and dataset library agree with the narrative/report
    # instead of showing a divergent 100 vs 30/D. Falls back to the
    # survival ratio if EDA scoring fails — must never break the clean run.
    quality_score = summary["quality_score"]
    quality_grade = None
    try:
        _dq = run_eda_auto(cleaned_df, effective_type).get("data_quality_score") or {}
        if _dq.get("score") is not None:
            quality_score = _dq["score"]
            quality_grade = _dq.get("grade")
    except Exception as exc:  # pragma: no cover - defensive
        logger.warning(
            "Quality rubric scoring failed for %s; using survival ratio: %s",
            filename,
            exc,
        )

    # Convert cleaned data to records
    cleaned_records = cleaned_df.replace({np.nan: None}).to_dict(orient="records")

    # Flag-then-filter: count of rows set aside (flagged, not deleted) so the UI
    # can show an honest denominator — "rates based on N of M records". The rows
    # themselves stay available via the full export and are summarised per-rule
    # in the Quality Report's "Records Dropped" tab.
    if "analyzable" in cleaned_df.columns:
        excluded_count = int((~cleaned_df["analyzable"]).sum())
    else:
        excluded_count = 0

    # Cache cleaned DF (durable) with enriched stats so downstream pages
    # (Explorer, KPI map, AI narrative) can resolve it by cache_id later.
    new_cache_id = _cache_cleaned(
        cleaned_df,
        {
            **(stats or {}),
            "filename": filename,
            "source_type": effective_type,
            "quality_score": quality_score,
            "quality_grade": quality_grade,
        },
    )

    persisted = True
    persist_error: Optional[str] = None
    try:
        _persist_session(
            cache_id=new_cache_id,
            filename=filename,
            source_type=effective_type,
            row_count=stats.get("final_count", len(cleaned_df)),
            result={
                **(stats or {}),
                "quality_score": quality_score,
                "issues": summary["top_issues"],
            },
            db=db,
            name=(body.dataset_name if body else None),
            owner=owner,
        )
    except Exception as exc:  # best-effort — never fail the clean run on a DB error
        persisted = False
        persist_error = f"{type(exc).__name__}: {exc}"
        logger.warning(
            "Session persistence failed for cache_id=%s (%s); the dataset will "
            "NOT appear on the dashboard until the database is reachable.",
            new_cache_id,
            persist_error,
        )

    # P2-2: Auto-persist child records to the durable linkage store — isolated
    # best-effort. A failure here must NEVER flip the session `persisted` flag
    # above (that flag reflects only whether the dataset will reach the dashboard).
    try:
        child_recs = _records_from_cached(
            new_cache_id, dataset_id=new_cache_id, source_type=effective_type
        )
        if child_recs:
            _persist_child_records(db, new_cache_id, effective_type, child_recs)
    except Exception as exc:  # noqa: BLE001 — best-effort, log and move on
        logger.warning(
            "Child-record persistence failed for cache_id=%s (%s: %s); durable "
            "linkage store not updated for this dataset.",
            new_cache_id,
            type(exc).__name__,
            exc,
        )

    # Build the full evaluated-rule set for this cleaner type so Quality
    # Report can show all checks including ones that passed (count=0).
    _rule_codes = EVALUATED_RULES.get(effective_type, EVALUATED_RULES["general"])
    rules_evaluated = [
        {
            "code": c,
            "count": int(stats.get(c, 0)),
            "fired": int(stats.get(c, 0)) > 0,
            "locked": c in LOCKED_RULES,
            "enabled": (enabled is None) or (c in LOCKED_RULES) or (c in enabled),
        }
        for c in _rule_codes
    ]

    _log_audit(action="clean.run", detail=f"cache_id={new_cache_id}", actor=owner)
    return JSONResponse(
        content=json_safe(
            {
                "success": True,
                "data_type": effective_type,
                "stats": stats,
                "cleaned_columns": cleaned_df.columns.tolist(),
                "cleaned_column_profile": _profile_columns(cleaned_df),
                "cleaned_count": stats.get("final_count", len(cleaned_df)),
                "preview": cleaned_records[:100],  # First 100 rows for preview
                "cache_id": new_cache_id,
                "rows_before": len(df),
                "rows_after": stats.get("final_count", len(cleaned_df)),
                "excluded_count": excluded_count,
                "quality_score": quality_score,
                "quality_grade": quality_grade,
                "rules_applied": summary["rules_applied"],
                "rules": summary["rules"],
                "top_issues": summary["top_issues"],
                "rules_evaluated": rules_evaluated,
                "persisted": persisted,
                "persist_error": persist_error,
            }
        )
    )


class PreviewImpactBody(BaseModel):
    """B3: proposed mapping + the rule codes the user kept enabled."""

    mapping: dict[str, str] = {}
    enabled_rules: Optional[List[str]] = None


def _resolve_effective_type(cache_id: Optional[str], data_type: str) -> str:
    """Same precedence as /clean/run: cached detected type wins over the default
    'myvass' query value; an explicit non-default caller value still overrides."""
    cached_st = ((_cache_get(cache_id) or {}).get("stats", {}) or {}).get("source_type")
    return cached_st if (data_type == "myvass" and cached_st) else data_type


@app.get("/clean/rules")
def clean_rules(data_type: str = Query("myvass")):
    """Registry view of drop + review rules for a source type, annotated with the
    user's persisted enabled state. Both groups share a vocabulary with Settings."""
    drop_state: dict = {}
    review_state: dict = {}
    try:
        from .db.init_db import SessionLocal as _SessionLocal

        if _SessionLocal is not None:
            _db = _SessionLocal()
            try:
                drop_state = _load_rule_state(_db)
                review_state = _load_review_rule_state(_db)
            finally:
                _db.close()
    except Exception:  # pragma: no cover - settings store unavailable
        drop_state = {}
        review_state = {}
    drop_rules = [
        {**r, "enabled": drop_state.get(r["code"], True), "kind": "drop"}
        for r in rules_for_source(data_type)
    ]
    review_rules = [
        {
            **r,
            "enabled": review_state.get(r["code"], True),
            "locked": False,
            "kind": "review",
        }
        for r in review_rules_for_source(data_type)
    ]
    return JSONResponse(
        content=json_safe({"data_type": data_type, "rules": drop_rules + review_rules})
    )


@app.post("/clean/preview-impact")
async def clean_preview_impact(
    cache_id: Optional[str] = Query(None),
    data_type: str = Query("myvass"),
    body: Optional[PreviewImpactBody] = None,
):
    """B3.2 live row-impact: run the real cleaner with the proposed enabled_rules
    and report the TRUE resulting row count + per-rule drops. Honest (no z-score
    short-cut that would overstate impact) and side-effect-free — never caches or
    persists. Re-run per toggle; debounce on the client."""
    df = _resolve_cached_df(cache_id)
    if body and getattr(body, "mapping", None):
        rename = {
            raw: std
            for raw, std in body.mapping.items()
            if std and raw in df.columns and raw != std
        }
        if rename:
            df = df.rename(columns=rename)
    effective_type = _resolve_effective_type(cache_id, data_type)
    enabled = (
        set(body.enabled_rules) if (body and body.enabled_rules is not None) else None
    )
    rows_before = len(df)
    try:
        cleaned_df, stats = clean_data(df, effective_type, enabled)
    except Exception as e:
        raise HTTPException(500, f"Preview error: {str(e)}")
    _rule_codes = EVALUATED_RULES.get(effective_type, EVALUATED_RULES["general"])
    per_rule = [
        {
            "code": c,
            "count": int(stats.get(c, 0)),
            "fired": int(stats.get(c, 0)) > 0,
            "locked": c in LOCKED_RULES,
            "enabled": (enabled is None) or (c in LOCKED_RULES) or (c in enabled),
        }
        for c in _rule_codes
    ]
    return JSONResponse(
        content=json_safe(
            {
                "rows_before": rows_before,
                "rows_after": stats.get("final_count", len(cleaned_df)),
                "per_rule": per_rule,
                "source_type": effective_type,
            }
        )
    )


@app.post("/clean/download")
async def clean_download_endpoint(
    file: UploadFile = File(...),
    data_type: str = "myvass",
    sheet: Optional[str] = None,
    fmt: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    view: str = Query("analysis", pattern="^(analysis|full)$"),
):
    """Download cleaned data as CSV or Excel.

    ``view=analysis`` (default) = analyzable rows only, flag columns dropped;
    ``view=full`` = every row plus ``analyzable``/``exclude_reason``.
    """
    content = await file.read()
    filename = file.filename or ""
    try:
        df, _ = read_file(content, filename, sheet)
    except Exception as e:
        raise HTTPException(400, str(e))

    try:
        cleaned_df, stats = clean_data(df, data_type)
    except Exception as e:
        raise HTTPException(500, f"Cleaning error: {str(e)}")

    if stats.get("final_count", len(cleaned_df)) == 0:
        raise HTTPException(422, "No data after cleaning")

    cleaned_df = _download_view(cleaned_df, view)
    base = filename.rsplit(".", 1)[0]
    timestamp = pd.Timestamp.now().strftime("%Y%m%d")

    if fmt == "xlsx":
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Cleaned data sheet
            cleaned_df.to_excel(writer, sheet_name=_view_sheet(view), index=False)

            # Stats sheet
            stats_df = pd.DataFrame(
                [
                    {"Metric": k, "Value": v}
                    for k, v in stats.items()
                    if not isinstance(v, dict)
                ]
            )
            stats_df.to_excel(writer, sheet_name="Cleaning Stats", index=False)

        output.seek(0)
        return StreamingResponse(
            iter([output.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{data_type.upper()}_{_view_label(view)}_{timestamp}.xlsx"'
            },
        )
    else:
        return StreamingResponse(
            _csv_stream(cleaned_df),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{data_type.upper()}_{_view_label(view)}_{timestamp}.csv"'
            },
        )


# ── Multi-file quality check ──────────────────────────────────────────────────
@app.post("/clean/quality-check-multi")
async def quality_check_multi_endpoint(
    files: List[UploadFile] = File(...),
    data_type: str = "myvass",
    ic_col: str = Query("IC_NO_PASSPORT"),
    dose_date_col: str = Query("DOSE_DATE"),
):
    """Merge multiple files, then return quality stats on the merged data."""
    file_contents = []
    for f in files:
        content = await f.read()
        file_contents.append((f.filename or "unknown", content))

    try:
        merged, merge_stats = _merge_myvass_files(file_contents, dose_date_col, ic_col)
    except ValueError as e:
        raise HTTPException(400, str(e))

    df = merged
    quality = {
        "total_rows": len(df),
        "total_columns": len(df.columns),
        "merge_stats": merge_stats,
        "columns": [],
    }

    for col in df.columns:
        col_data = df[col]
        non_null = col_data.notna().sum()
        null_count = len(df) - non_null
        unique_count = col_data.nunique()
        numeric_vals = pd.to_numeric(col_data, errors="coerce")
        is_numeric = numeric_vals.notna().sum() > non_null * 0.8

        col_info = {
            "name": col,
            "non_null": int(non_null),
            "null_count": int(null_count),
            "null_percent": round((null_count / len(df)) * 100, 1)
            if len(df) > 0
            else 0,
            "unique_count": int(unique_count),
            "is_numeric": is_numeric,
        }
        if is_numeric and non_null > 0:
            col_info["min"] = (
                float(numeric_vals.min()) if numeric_vals.notna().any() else None
            )
            col_info["max"] = (
                float(numeric_vals.max()) if numeric_vals.notna().any() else None
            )
            col_info["mean"] = (
                round(float(numeric_vals.mean()), 2)
                if numeric_vals.notna().any()
                else None
            )
        else:
            sample = col_data.dropna().head(5).tolist()
            col_info["sample_values"] = [str(v)[:50] for v in sample]
        quality["columns"].append(col_info)

    total_cells = len(df) * len(df.columns)
    filled_cells = sum(c["non_null"] for c in quality["columns"])
    quality["overall_completeness"] = (
        round((filled_cells / total_cells) * 100, 1) if total_cells > 0 else 0
    )

    return JSONResponse(content=json_safe(quality))


# ── Multi-file cleaning ───────────────────────────────────────────────────────
@app.post("/clean/run-multi")
async def clean_run_multi_endpoint(
    files: List[UploadFile] = File(...),
    data_type: str = "myvass",
    ic_col: str = Query("IC_NO_PASSPORT"),
    dose_date_col: str = Query("DOSE_DATE"),
):
    """Merge multiple files, then run cleaning on the merged data."""
    file_contents = []
    for f in files:
        content = await f.read()
        file_contents.append((f.filename or "unknown", content))

    try:
        merged, merge_stats = _merge_myvass_files(file_contents, dose_date_col, ic_col)
    except ValueError as e:
        raise HTTPException(400, str(e))

    try:
        cleaned_df, stats = clean_data(merged, data_type)
    except Exception as e:
        raise HTTPException(500, f"Cleaning error: {str(e)}")

    stats["merge_stats"] = merge_stats
    cleaned_records = cleaned_df.replace({np.nan: None}).to_dict(orient="records")

    # Cache cleaned DF so download doesn't need re-upload
    cache_id = _cache_cleaned(cleaned_df, stats)

    return JSONResponse(
        content=json_safe(
            {
                "success": True,
                "data_type": data_type,
                "stats": stats,
                "cleaned_columns": cleaned_df.columns.tolist(),
                "cleaned_column_profile": _profile_columns(cleaned_df),
                "cleaned_count": stats.get("final_count", len(cleaned_df)),
                "preview": cleaned_records[:100],
                "cache_id": cache_id,
            }
        )
    )


# ── Multi-file download ───────────────────────────────────────────────────────
@app.post("/clean/download-multi")
async def clean_download_multi_endpoint(
    files: List[UploadFile] = File(...),
    data_type: str = "myvass",
    ic_col: str = Query("IC_NO_PASSPORT"),
    dose_date_col: str = Query("DOSE_DATE"),
    fmt: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    view: str = Query("analysis", pattern="^(analysis|full)$"),
):
    """Merge multiple files, clean, and download.

    ``view`` follows the same analysis/full contract as ``/clean/download``.
    """
    file_contents = []
    for f in files:
        content = await f.read()
        file_contents.append((f.filename or "unknown", content))

    try:
        merged, merge_stats = _merge_myvass_files(file_contents, dose_date_col, ic_col)
    except ValueError as e:
        raise HTTPException(400, str(e))

    try:
        cleaned_df, stats = clean_data(merged, data_type)
    except Exception as e:
        raise HTTPException(500, f"Cleaning error: {str(e)}")

    if stats.get("final_count", len(cleaned_df)) == 0:
        raise HTTPException(422, "No data after cleaning")

    cleaned_df = _download_view(cleaned_df, view)
    timestamp = pd.Timestamp.now().strftime("%Y%m%d")

    if fmt == "xlsx":
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            cleaned_df.to_excel(writer, sheet_name=_view_sheet(view), index=False)
            stats_rows = [
                {"Metric": k, "Value": v}
                for k, v in stats.items()
                if not isinstance(v, dict)
            ]
            merge_rows = [
                {"Metric": f"merge_{k}", "Value": v}
                for k, v in merge_stats.items()
                if not isinstance(v, (dict, list))
            ]
            pd.DataFrame(stats_rows + merge_rows).to_excel(
                writer, sheet_name="Cleaning Stats", index=False
            )
        output.seek(0)
        return StreamingResponse(
            iter([output.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{data_type.upper()}_Merged_{_view_label(view)}_{timestamp}.xlsx"'
            },
        )
    else:
        return StreamingResponse(
            _csv_stream(cleaned_df),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{data_type.upper()}_Merged_{_view_label(view)}_{timestamp}.csv"'
            },
        )


# ── Cached download (instant — no re-upload) ─────────────────────────────────
@app.get("/clean/download-cached/{cache_id}")
async def download_cached_endpoint(
    cache_id: str,
    fmt: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    data_type: str = Query("myvass"),
    view: str = Query("analysis", pattern="^(analysis|full)$"),
):
    """Download previously cleaned data from cache — no file re-upload needed.

    ``view=analysis`` (default) returns analyzable rows only, without the
    internal flag columns — what users expect when they ask for "cleaned data".
    ``view=full`` returns every row plus ``analyzable``/``exclude_reason`` so an
    analyst can audit exactly which rows were excluded and why.
    """
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(404, "Cached data not found — please re-run cleaning.")
    cleaned_df = _download_view(entry["df"], view)

    timestamp = pd.Timestamp.now().strftime("%Y%m%d")

    if fmt == "xlsx":
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            cleaned_df.to_excel(writer, sheet_name=_view_sheet(view), index=False)
        output.seek(0)
        return StreamingResponse(
            iter([output.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{data_type.upper()}_{_view_label(view)}_{timestamp}.xlsx"'
            },
        )
    else:
        return StreamingResponse(
            _csv_stream(cleaned_df),
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{data_type.upper()}_{_view_label(view)}_{timestamp}.csv"'
            },
        )


# ── Cached JSON preview (for the Explorer page) ──────────────────────────────
@app.get("/clean/preview-cached/{cache_id}")
async def preview_cached_endpoint(
    cache_id: str,
    limit: int = Query(500, ge=1, le=5000),
):
    """Return rows from the cached cleaned dataset as JSON for the Explorer.

    The Explorer no longer depends on transient in-memory frontend state — it
    fetches by cache_id, which is durable across restarts via the disk cache.
    """
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(404, "Cached data not found — please re-run cleaning.")
    df = _explorer_view(entry["df"])
    head = df.head(limit)
    rows = head.replace({np.nan: None}).to_dict(orient="records")
    row_flags = _compute_row_flags(head).tolist()
    return JSONResponse(
        content=json_safe(
            {
                "columns": df.columns.tolist(),
                "rows": rows,
                "row_count": int(len(df)),
                "returned": len(rows),
                "row_flags": row_flags,
            }
        )
    )


# ── Stable-key query seam (Phase 5 — edit-under-filter + future scale) ────────
class QueryCachedRequest(BaseModel):
    cache_id: str
    search: str | None = None
    sort_col: str | None = None
    sort_dir: str = "asc"  # "asc" | "desc"
    offset: int = Field(0, ge=0)
    limit: int = Field(10000, ge=1, le=100000)


@app.post("/clean/query-cached")
async def query_cached_seam(req: QueryCachedRequest):
    """Full-dataset query endpoint with stable _row_id per row.

    _row_id = positional iloc index in the original (cleaned) DataFrame.
    Edits via PATCH /clean/cell route by _row_id — safe under any sort/filter.
    Phase 7 can swap the pandas backend to DuckDB/Parquet behind this contract.
    """
    entry = _cache_get(req.cache_id)
    if entry is None:
        raise HTTPException(404, "cache_id not found — please re-run cleaning.")

    df = _explorer_view(entry["df"]).reset_index(drop=True)
    flags = _compute_row_flags(df)

    # Embed stable ids and per-row flag before any filter/sort
    result = df.copy()
    result.insert(0, "_row_id", range(len(result)))
    result.insert(1, "_flagged", flags.values.astype(bool))

    # Apply search across all non-private columns
    if req.search:
        q = req.search.lower()
        data_cols = [c for c in result.columns if not c.startswith("_")]
        mask = (
            result[data_cols]
            .astype(str)
            .apply(
                lambda row: any(
                    q in v.lower() for v in row if v.lower() not in ("nan", "none")
                ),
                axis=1,
            )
        )
        result = result[mask]

    # Apply sort (never sort on _row_id / _flagged themselves)
    if (
        req.sort_col
        and req.sort_col in result.columns
        and not req.sort_col.startswith("_")
    ):
        result = result.sort_values(
            req.sort_col,
            ascending=(req.sort_dir != "desc"),
            na_position="last",
        )

    total = int(len(result))
    window = result.iloc[req.offset : req.offset + req.limit]
    rows = window.replace({np.nan: None}).to_dict(orient="records")

    return JSONResponse(
        content=json_safe(
            {
                "rows": rows,
                "total": total,
                "returned": len(rows),
                "offset": req.offset,
            }
        )
    )


@app.get("/clean/download-xlsx/{cache_id}")
async def download_xlsx_endpoint(
    cache_id: str,
    view: str = Query("analysis", pattern="^(analysis|full)$"),
):
    """Download the cleaned dataset as XLSX with IC/text columns typed correctly.

    Unlike the generic download-cached endpoint, this forces object-dtype and
    IC-keyword columns to Excel text format (@) so leading zeros in MyKid/IC
    numbers are preserved when the file is opened in Excel.

    ``view`` follows the same analysis/full contract as ``download-cached``.
    """
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(404, "Cached data not found — please re-run cleaning.")

    df = _download_view(entry["df"], view)
    timestamp = pd.Timestamp.now().strftime("%Y%m%d")
    xlsx_bytes = cln_excel_typed(df)

    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="SmartDQC_{_view_label(view)}_{timestamp}.xlsx"'
        },
    )


def _cache_write(key: str, entry: dict) -> None:
    """Persist a cache entry to the hot tier and disk so edits survive restart."""
    _cleaned_cache[key] = entry
    try:
        with _cache_path(key).open("wb") as fh:
            _pickle.dump(entry, fh, protocol=_pickle.HIGHEST_PROTOCOL)
    except Exception as exc:  # pragma: no cover - best effort
        logger.warning("Cache write failed for %s: %s", key, exc)


class CellEditRequest(BaseModel):
    cache_id: str
    row_index: int
    column: str
    value: object


@app.patch("/clean/cell")
async def edit_cell(req: CellEditRequest):
    """Edit a single cell of the cached cleaned dataset (in-memory + disk, audit-logged)."""
    entry = _cache_get(req.cache_id)
    if entry is None:
        raise HTTPException(404, "cache_id not found — re-run cleaning first.")
    df = entry["df"]
    if req.column not in df.columns:
        raise HTTPException(400, f"Column '{req.column}' not in dataset.")
    if req.row_index < 0 or req.row_index >= len(df):
        raise HTTPException(
            400, f"row_index {req.row_index} out of range (0..{len(df) - 1})."
        )

    old = df.iloc[req.row_index][req.column]
    # Coerce to the column's dtype where possible; fall back to raw value.
    new_val = req.value
    try:
        if pd.api.types.is_numeric_dtype(df[req.column]):
            new_val = pd.to_numeric(req.value)
    except (ValueError, TypeError):
        new_val = req.value
    df.iloc[req.row_index, df.columns.get_loc(req.column)] = new_val

    entry["df"] = df
    _cache_write(req.cache_id, entry)
    _log_audit(
        action="clean.cell_edit",
        detail=f"{req.cache_id} row={req.row_index} col={req.column} {old!r}->{new_val!r}",
    )
    updated_row = (
        df.iloc[[req.row_index]].replace({np.nan: None}).to_dict(orient="records")[0]
    )
    return JSONResponse(
        content=json_safe({"row_index": req.row_index, "row": updated_row})
    )


# ── Dataset Join (horizontal / vertical union) ───────────────────────────────

_JOIN_TYPES = {"inner", "left", "right", "outer", "union"}


@app.post("/join/preview")
async def join_preview_endpoint(
    file_left: UploadFile | None = File(None),
    file_right: UploadFile | None = File(None),
    cache_id_left: str | None = Query(None),
    cache_id_right: str | None = Query(None),
    join_type: str = Query(..., pattern="^(inner|left|right|outer|union)$"),
    key_cols: str | None = Query(
        None, description="Comma-separated key column names (horizontal joins)"
    ),
    dedup: bool = Query(False, description="Remove duplicate rows after union"),
):
    """Preview a join of two datasets. Returns first 50 rows plus shape and stats."""
    left_bytes = (await file_left.read()) if file_left else None
    right_bytes = (await file_right.read()) if file_right else None

    df_left = _resolve_source(
        left_bytes, file_left.filename if file_left else None, cache_id_left
    )
    df_right = _resolve_source(
        right_bytes, file_right.filename if file_right else None, cache_id_right
    )

    parsed_keys = [c.strip() for c in key_cols.split(",")] if key_cols else None
    result, stats = _perform_join(df_left, df_right, join_type, parsed_keys, dedup)

    return {
        "preview": result.head(50).to_dict(orient="records"),
        "columns": list(result.columns),
        "left_columns": list(df_left.columns),
        "right_columns": list(df_right.columns),
        "shape": {"rows": len(result), "cols": len(result.columns)},
        "left_shape": {"rows": len(df_left), "cols": len(df_left.columns)},
        "right_shape": {"rows": len(df_right), "cols": len(df_right.columns)},
        "join_stats": stats,
    }


def _join_source_label(cache_id: str | None, upload_filename: str | None) -> str:
    """Resolve a human-friendly label for a join side: prefer the cached
    dataset filename, fall back to the upload filename, then to the
    truncated cache_id."""
    if cache_id:
        entry = _cache_get(cache_id)
        stats = (entry or {}).get("stats") or {}
        return stats.get("filename") or cache_id[:8]
    return upload_filename or "upload"


@app.post("/join/run")
async def join_run_endpoint(
    file_left: UploadFile | None = File(None),
    file_right: UploadFile | None = File(None),
    cache_id_left: str | None = Query(None),
    cache_id_right: str | None = Query(None),
    join_type: str = Query(..., pattern="^(inner|left|right|outer|union)$"),
    key_cols: str | None = Query(
        None, description="Comma-separated key column names (horizontal joins)"
    ),
    dedup: bool = Query(False, description="Remove duplicate rows after union"),
    owner: str | None = Depends(_identity),
    db=Depends(get_db),
):
    """Execute a full join, cache the result, and persist it as a library Dataset row."""
    left_bytes = (await file_left.read()) if file_left else None
    right_bytes = (await file_right.read()) if file_right else None

    df_left = _resolve_source(
        left_bytes, file_left.filename if file_left else None, cache_id_left
    )
    df_right = _resolve_source(
        right_bytes, file_right.filename if file_right else None, cache_id_right
    )

    parsed_keys = [c.strip() for c in key_cols.split(",")] if key_cols else None
    result, stats = _perform_join(df_left, df_right, join_type, parsed_keys, dedup)

    # Synthesise a human-friendly filename for the joined dataset.
    left_label = _join_source_label(
        cache_id_left, file_left.filename if file_left else None
    )
    right_label = _join_source_label(
        cache_id_right, file_right.filename if file_right else None
    )
    join_symbol = "∪" if join_type == "union" else "⨝"
    joined_name = f"{left_label} {join_symbol} {right_label} ({join_type})"

    # Inherit source_type from the left side when both sides agree, so the
    # cleaner downstream knows what shape to expect. Otherwise mark as
    # "joined" — the cleaner will fall through to the generic path.
    left_st = (
        ((_cache_get(cache_id_left) or {}).get("stats") or {}).get("source_type")
        if cache_id_left
        else None
    )
    right_st = (
        ((_cache_get(cache_id_right) or {}).get("stats") or {}).get("source_type")
        if cache_id_right
        else None
    )
    effective_st = left_st if (left_st and left_st == right_st) else "joined"

    cache_id = _cache_cleaned(
        result,
        {
            "filename": joined_name,
            "source_type": effective_st,
            "rows": len(result),
            "cols": len(result.columns),
            "join_type": join_type,
            "join_stats": stats,
        },
    )

    # Best-effort persist so the joined dataset shows up in the library /
    # history list. A DB failure here must not break the join.
    try:
        _persist_session(
            cache_id=cache_id,
            filename=joined_name,
            source_type=effective_st,
            row_count=len(result),
            result={"join_type": join_type, "join_stats": stats},
            db=db,
            owner=owner,
        )
    except Exception as exc:  # pragma: no cover — defensive
        logger.warning("Joined-dataset persist failed for %s: %s", cache_id, exc)

    _log_audit(
        action="dataset.join",
        dataset_id=cache_id,
        detail=f"{join_type}: {left_label} + {right_label}",
    )

    return {
        "cache_id": cache_id,
        "shape": {"rows": len(result), "cols": len(result.columns)},
        "join_stats": stats,
    }


# --- ML NAMESPACE ---------------------------------------------------------------


@app.post("/ml/suggest")
async def ml_suggest_endpoint(
    cache_id: str = Query(..., description="UUID from /clean/run or /join/run"),
):
    """Flag anomalous rows and suggest corrections using IsolationForest."""
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(
            404, "cache_id not found — run /clean/run first or check the UUID"
        )
    result = flag_anomalies(entry["df"])
    return JSONResponse(content=json_safe(result))


# --- REPORT NAMESPACE -----------------------------------------------------------


def _parse_charts(charts: str | None) -> set[str] | None:
    """Parse a comma-separated charts query into a set of keys.

    - param absent/empty  → None  (caller wants the default recommended set)
    - "none" sentinel     → empty set (caller explicitly wants no charts)
    - otherwise           → the explicit set of keys
    """
    if not charts:
        return None
    keys = {c.strip() for c in charts.split(",") if c.strip()}
    keys.discard("none")
    return keys


@app.get("/report/pptx")
def report_pptx_endpoint(
    cache_id: str = Query(..., description="UUID from /clean/run or /join/run"),
    include_kpi: bool = Query(True, description="Embed KPI dashboard slides"),
    charts: str | None = Query(
        None,
        description=(
            "Comma-separated chart keys to embed; omit for the recommended "
            "defaults. Known keys: quality_bar, nutritional_rates, kpi_vs_target."
        ),
    ),
    db=Depends(get_db),
):
    """Generate a PPTX report from the cached EDA result."""
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(404, "cache_id not found — run /clean/run first")
    # The cached entry["stats"] is the cleaner's per-rule counts, NOT a
    # run_eda result — feeding it to the report builder left the data
    # quality overview / indicators / charts blank. Build the real EDA
    # report (same run_eda_auto the AI narrative uses) so the report is
    # populated and consistent with the narrative + dashboard score.
    _src = (entry.get("stats") or {}).get("source_type") or "myvass"
    eda_result = run_eda_auto(entry["df"], _src)
    _tgt = _load_kpi_targets(db)
    _amber, _ = _rag_tolerances(db)
    kpi_result = (
        compute_kpi_dashboard(
            entry["df"], npan=_tgt["npan"], who=_tgt["who"], amber_tolerance=_amber
        )
        if include_kpi
        else None
    )
    narrative = _get_or_build_narrative(cache_id, entry)
    data = build_pptx_bytes(
        eda_result, narrative, kpi_result=kpi_result, charts=_parse_charts(charts)
    )
    _log_audit(
        action="report.pptx", detail=f"cache_id={cache_id} charts={charts or 'default'}"
    )
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": 'attachment; filename="SmartDQC_Report.pptx"'},
    )


@app.get("/report/pdf")
def report_pdf_endpoint(
    cache_id: str = Query(..., description="UUID from /clean/run or /join/run"),
    include_kpi: bool = Query(True, description="Embed KPI dashboard section"),
    charts: str | None = Query(
        None,
        description=(
            "Comma-separated chart keys to embed; omit for the recommended "
            "defaults. Known keys: quality_bar, nutritional_rates, kpi_vs_target."
        ),
    ),
    db=Depends(get_db),
):
    """Generate a PDF report from the cached EDA result."""
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(404, "cache_id not found — run /clean/run first")
    # The cached entry["stats"] is the cleaner's per-rule counts, NOT a
    # run_eda result — feeding it to the report builder left the data
    # quality overview / indicators / charts blank. Build the real EDA
    # report (same run_eda_auto the AI narrative uses) so the report is
    # populated and consistent with the narrative + dashboard score.
    _src = (entry.get("stats") or {}).get("source_type") or "myvass"
    eda_result = run_eda_auto(entry["df"], _src)
    _tgt = _load_kpi_targets(db)
    _amber, _ = _rag_tolerances(db)
    kpi_result = (
        compute_kpi_dashboard(
            entry["df"], npan=_tgt["npan"], who=_tgt["who"], amber_tolerance=_amber
        )
        if include_kpi
        else None
    )
    narrative = _get_or_build_narrative(cache_id, entry)
    data = build_pdf_bytes(
        eda_result, narrative, kpi_result=kpi_result, charts=_parse_charts(charts)
    )
    _log_audit(
        action="report.pdf", detail=f"cache_id={cache_id} charts={charts or 'default'}"
    )
    return Response(
        content=data,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="SmartDQC_Report.pdf"'},
    )


# --- RISK NAMESPACE -------------------------------------------------------------


@app.post("/risk/score")
async def risk_score_endpoint(
    cache_id: str = Query(..., description="UUID from /clean/run or /join/run"),
):
    """Compute per-child composite risk score (0-100) and district aggregation."""
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(
            404, "cache_id not found — run /clean/run first or check the UUID"
        )
    result = compute_risk_scores(entry["df"])
    return JSONResponse(content=json_safe(result))


# --- KPI NAMESPACE --------------------------------------------------------------


@app.post("/kpi/dashboard")
async def kpi_dashboard_endpoint(
    cache_id: str = Query(..., description="UUID from /clean/run or /join/run"),
    district: str | None = Query(None, description="Filter by district name"),
    state: str | None = Query(
        None, description="Filter by state name (full name, case-insensitive)"
    ),
    db=Depends(get_db),
):
    """Return RAG traffic-light KPI status benchmarked against Malaysian national targets."""
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(
            404, "cache_id not found — run /clean/run first or check the UUID"
        )
    df = entry["df"]
    if district:
        district_col = next(
            (
                c
                for c in df.columns
                if any(k in c.lower() for k in ("district", "daerah", "kawasan"))
            ),
            None,
        )
        if district_col:
            df = df[df[district_col].str.lower() == district.lower()]
    if state:
        state_col = next(
            (c for c in df.columns if c.lower() in ("negeri", "state")),
            None,
        )
        if state_col:
            df = df[
                df[state_col].astype(str).str.strip().str.lower()
                == state.strip().lower()
            ]
    _tgt = _load_kpi_targets(db)
    _amber, _ = _rag_tolerances(db)
    result = compute_kpi_dashboard(
        df, npan=_tgt["npan"], who=_tgt["who"], amber_tolerance=_amber
    )
    return JSONResponse(content=json_safe(result))


@app.get("/charts/blocks")
async def charts_blocks_endpoint(
    cache_id: str = Query(..., description="UUID from /clean/run or /join/run"),
):
    """Return per-column histograms / scatter blocks for the cached dataset.

    Drives the "Data Distributions" section on GeoPage and the optional
    chart embeds in PDF/PPTX reports. Output is the same shape produced by
    build_chart_blocks(df) and embedded under "charts" inside run_eda's
    report — exposed here so already-cached datasets can render the
    blocks without re-running EDA.
    """
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(
            404, "cache_id not found — run /clean/run first or check the UUID"
        )
    try:
        blocks = build_chart_blocks(entry["df"])
    except Exception as exc:  # pragma: no cover — defensive
        logger.warning("build_chart_blocks failed for %s: %s", cache_id, exc)
        raise HTTPException(500, f"Chart computation failed: {exc}")
    return JSONResponse(content=json_safe(blocks))


@app.get("/quality/breakdown")
async def quality_breakdown_endpoint(
    cache_id: str = Query(..., description="UUID from /clean/run or /join/run"),
):
    """Return the 7-dimension quality-score breakdown for a cached dataset.

    Exposes the data_quality_score block that compute_quality_score already
    produces inside run_eda_auto, so the Quality page can render a
    by-dimension breakdown by cache_id (including for reopened sessions)
    without the breakdown having to ride along in every /clean/run response.
    Mirrors the /charts/blocks recompute-from-cache pattern.
    """
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(
            404, "cache_id not found — run /clean/run first or check the UUID"
        )
    source_type = ((entry.get("stats") or {}).get("source_type")) or "myvass"
    try:
        dq = run_eda_auto(entry["df"], source_type).get("data_quality_score") or {}
    except Exception as exc:  # pragma: no cover — defensive
        logger.warning("Quality breakdown computation failed for %s: %s", cache_id, exc)
        raise HTTPException(500, f"Quality breakdown computation failed: {exc}")
    return JSONResponse(content=json_safe(dq))


class TrajectoryRequest(BaseModel):
    historical_snapshots: list[dict]
    current_breakdown: list[dict] = []


@app.post("/kpi/trajectory")
async def kpi_trajectory(req: TrajectoryRequest, db=Depends(get_db)):
    """Compute per-district trajectory narratives and 2027 target forecast from indicator snapshots."""
    return compute_trajectory_narratives(
        req.historical_snapshots,
        req.current_breakdown,
        npan=_load_kpi_targets(db)["npan"],
        atrisk_tolerance=_rag_tolerances(db)[1],
        target_year=_load_target_year(db),
    )


@app.post("/kpi/trajectory/auto")
async def kpi_trajectory_auto(
    cache_id: str = Query(..., description="UUID from /clean/run or /join/run"),
    db=Depends(get_db),
):
    """Per-district trajectory narratives derived from the cached dataset's own
    multi-year (tahun_ukur) data. Returns an empty narrative list for single-year
    datasets — trajectory needs >=2 measurement periods per district."""
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(
            404, "cache_id not found — run /clean/run first or check the UUID"
        )
    snapshots = compute_district_period_snapshots(entry["df"])
    narratives = compute_trajectory_narratives(
        snapshots,
        [],
        npan=_load_kpi_targets(db)["npan"],
        atrisk_tolerance=_rag_tolerances(db)[1],
        target_year=_load_target_year(db),
    )
    periods = sorted({s["period"] for s in snapshots})
    return JSONResponse(
        content=json_safe(
            {
                "narratives": narratives,
                "periods": periods,
                "has_multiyear": len(periods) >= 2,
            }
        )
    )


# ── Data Quality Report (5-tab Excel) ────────────────────────────────────────


def _build_quality_report(df: pd.DataFrame, stats: dict, data_type: str) -> io.BytesIO:
    """Build a multi-tab Excel data quality report from cleaned data + cleaning stats."""
    from datetime import datetime
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    raw = stats.get("raw_count", 0)
    final = stats.get("final_count", len(df))

    def pct(n, base):
        return f"{n / base * 100:.2f}%" if base else "0%"

    def _style_sheet(ws, header_row=1, freeze_row=2):
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for cell in ws[header_row]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 4, 45)
        ws.freeze_panes = ws.cell(row=freeze_row, column=1)

    # Rule-level stats
    dropped_gender = stats.get("dropped_invalid_gender", 0)
    dropped_dob = stats.get("dropped_date_before_dob", 0)
    dropped_age = stats.get("dropped_age_over5", stats.get("dropped_age_invalid", 0))
    dropped_outlier = stats.get("dropped_measurement_outlier", 0)
    dropped_bmi = stats.get("dropped_bmi_outlier", 0)
    dropped_no_meas = stats.get("dropped_no_measurement", 0)
    dropped_zscore = stats.get("dropped_null_zscore", 0)

    rules_ordered = [
        ("Rule 1", "Jantina 'Others'", dropped_gender),
        ("Rule 4", "Assessment before DOB", dropped_dob),
        ("Rule 3", "Age > 5 years", dropped_age),
        ("Rule 2", "Measurement outliers", dropped_outlier),
        ("Rule 5", "Implausible BMI > 40", dropped_bmi),
        ("Rule 6", "No measurement (both null)", dropped_no_meas),
        ("Rule 7", "Null z-score(s)", dropped_zscore),
    ]

    # TAB 1 — Executive Summary
    exec_rows = [
        [f"{data_type.upper()} \u2014 DATA QUALITY REPORT", "", "", ""],
        [f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}", "", "", ""],
        [],
        ["Metric", "Count", "% of Raw", "Notes"],
        ["Raw records (before cleaning)", raw, "100%", "Source CSV"],
    ]
    for rule_id, label, count in rules_ordered:
        exec_rows.append([f"Dropped \u2014 {label}", count, pct(count, raw), rule_id])
    exec_rows.append(
        ["Final cleaned records", final, pct(final, raw), "After all rules"]
    )
    exec_rows.append([])

    # WAZ distribution
    exec_rows.append(["WAZ \u2014 Weight-for-Age", "Count", "% of Cleaned", ""])
    if "WAZ_Status" in df.columns:
        waz_vc = df["WAZ_Status"].value_counts()
        for code, label in [
            ("berat_badan_normal", "Berat Badan Normal"),
            ("risiko_kurang_berat_badan", "Risiko Kurang Berat Badan"),
            ("kurang_berat_badan", "Kurang Berat Badan"),
            ("kurang_berat_badan_teruk", "Kurang Berat Badan Teruk"),
            ("mungkin_masalah_pertumbuhan", "Mungkin Masalah Pertumbuhan"),
        ]:
            cnt = int(waz_vc.get(code, 0))
            exec_rows.append([label, cnt, pct(cnt, final), ""])
    exec_rows.append([])

    # HAZ distribution
    exec_rows.append(["HAZ \u2014 Height-for-Age", "Count", "% of Cleaned", ""])
    if "HAZ_Status" in df.columns:
        haz_vc = df["HAZ_Status"].value_counts()
        for code, label in [
            ("normal", "Panjang/Tinggi Normal"),
            ("risiko_bantut", "Risiko Bantut"),
            ("bantut", "Bantut"),
            ("bantut_teruk", "Bantut Teruk"),
            ("mungkin_masalah_endokrin", "Mungkin Masalah Endokrin"),
        ]:
            cnt = int(haz_vc.get(code, 0))
            exec_rows.append([label, cnt, pct(cnt, final), ""])
    exec_rows.append([])

    # BAZ distribution
    exec_rows.append(["BAZ \u2014 BMI-for-Age", "Count", "% of Cleaned", ""])
    if "BAZ_Status" in df.columns:
        baz_vc = df["BAZ_Status"].value_counts()
        for code, label in [
            ("normal", "Berat Badan Normal"),
            ("berisiko_susut", "Berisiko Susut"),
            ("susut", "Susut"),
            ("susut_teruk", "Susut Teruk"),
            ("risiko_lebih_berat_badan", "Risiko Lebih Berat Badan"),
            ("berlebihan_berat_badan", "Berlebihan Berat Badan"),
            ("obes", "Obes"),
        ]:
            cnt = int(baz_vc.get(code, 0))
            exec_rows.append([label, cnt, pct(cnt, final), ""])
    exec_rows.append([])
    if "Kategori_Umur" in df.columns:
        bawah2 = int((df["Kategori_Umur"] == "Bawah 2 Tahun").sum())
        bawah5 = int((df["Kategori_Umur"] == "Bawah 5 Tahun").sum())
    elif "Age_Months" in df.columns:
        age_m = pd.to_numeric(df["Age_Months"], errors="coerce")
        bawah2 = int((age_m < 24).sum())
        bawah5 = int(((age_m >= 24) & (age_m < 60)).sum())
    else:
        bawah2, bawah5 = 0, final
    exec_rows.append(["Age Group", "Count", "% of Cleaned", "Definition"])
    exec_rows.append(["Bawah 2 Tahun", bawah2, pct(bawah2, final), "0 \u2013 23 bulan"])
    exec_rows.append(
        ["Bawah 5 Tahun", bawah5, pct(bawah5, final), "24 \u2013 59 bulan"]
    )

    # TAB 2 — Cleaning Rules
    rules_detail = [
        ["CLEANING RULES APPLIED", "", "", "", "", ""],
        [],
        ["Rule", "Column(s)", "Condition", "Action", "Rows Affected", "% of Raw"],
        [
            "Rule 1",
            "Jantina",
            "Jantina = 'Others'",
            "Drop row",
            dropped_gender,
            pct(dropped_gender, raw),
        ],
        [
            "Rule 2",
            "Berat / Panjang",
            "Berat \u2264 0.5 or > 35.0 kg | Panjang \u2264 30.0 or > 130.0 cm",
            "Drop row",
            dropped_outlier,
            pct(dropped_outlier, raw),
        ],
        [
            "Rule 3",
            "Tarikh Lahir / Tarikh Antropometri",
            "Age at assessment \u2265 60 months (> 5 years)",
            "Drop row",
            dropped_age,
            pct(dropped_age, raw),
        ],
        [
            "Rule 4",
            "Tarikh Lahir / Tarikh Antropometri",
            "Tarikh Antropometri < Tarikh Lahir",
            "Drop row",
            dropped_dob,
            pct(dropped_dob, raw),
        ],
        [
            "Rule 5",
            "Berat + Panjang",
            "BMI > 40.0 (implausible weight/height combination)",
            "Drop row",
            dropped_bmi,
            pct(dropped_bmi, raw),
        ],
        [
            "Rule 6",
            "Berat + Panjang",
            "Both weight AND height are NULL",
            "Drop row",
            dropped_no_meas,
            pct(dropped_no_meas, raw),
        ],
        [
            "Rule 7",
            "WAZ / HAZ / BAZ",
            "Any z-score is NULL after computation",
            "Drop row",
            dropped_zscore,
            pct(dropped_zscore, raw),
        ],
    ]

    # TAB 3 — Records Dropped
    dropped_rows = [
        ["RECORDS DROPPED \u2014 Summary by Rule", "", ""],
        [],
        ["Rule", "Records Dropped", "Cumulative Remaining"],
    ]
    cumulative = raw
    for rule_id, label, count in rules_ordered:
        if count > 0:
            cumulative -= count
            dropped_rows.append([f"{rule_id} \u2014 {label}", count, cumulative])

    # ── Find geographic columns ────────────────────────────────────────────────
    state_col = None
    for c in df.columns:
        if c.lower() in ("state", "negeri"):
            state_col = c
            break
    if state_col is None:
        for c in df.columns:
            if "negeri" in c.lower() or "state" in c.lower():
                state_col = c
                break

    district_col = None
    for c in df.columns:
        if c.lower() in ("district", "daerah"):
            district_col = c
            break
    if district_col is None:
        for c in df.columns:
            if "daerah" in c.lower() or "district" in c.lower():
                district_col = c
                break

    # ── Helper: build pivot table (4 sections) for one z-score indicator ──────
    age_cats = ["Bawah 2 Tahun", "Bawah 5 Tahun"]

    def _build_pivot_tab(
        geo_col, geo_label, status_col, detail_labels, combine_map, combine_order, title
    ):
        """
        Build rows for one pivot tab.
        detail_labels: ordered list of classification codes
        combine_map: dict mapping combined_label -> list of codes to merge
        combine_order: ordered list of combined labels
        """
        rows_out = []
        if geo_col is None or geo_col not in df.columns or status_col not in df.columns:
            rows_out.append([f"(No {geo_label} or {status_col} column found)"])
            return rows_out

        # Ensure Kategori_Umur exists
        if "Kategori_Umur" not in df.columns:
            return [[f"(No Kategori_Umur column found)"]]

        # --- SECTION 1: Detailed Count ---
        hdr1 = [f"Count of {status_col}", "Column Labels"]
        for lbl in detail_labels:
            hdr1.extend(["", "", ""])
        rows_out.append(hdr1)

        sub_hdr1_1 = [f"", ""]
        for lbl in detail_labels:
            sub_hdr1_1.extend([lbl, "", f"{lbl} Total"])
        rows_out.append(sub_hdr1_1)

        sub_hdr1_2 = ["Row Labels", ""]
        for lbl in detail_labels:
            sub_hdr1_2.extend(["Bawah 2 Tahun", "Bawah 5 Tahun", ""])
        rows_out.append(sub_hdr1_2)

        # Group data
        grouped = (
            df.groupby([geo_col, "Kategori_Umur", status_col])
            .size()
            .reset_index(name="count")
        )
        geo_totals = df.groupby(geo_col).size()

        detail_data = []
        for geo_name in sorted(df[geo_col].dropna().unique()):
            row = [str(geo_name), int(geo_totals.get(geo_name, 0))]
            grp = grouped[grouped[geo_col] == geo_name]
            for lbl in detail_labels:
                for age_cat in age_cats:
                    mask = (grp["Kategori_Umur"] == age_cat) & (grp[status_col] == lbl)
                    val = int(grp.loc[mask, "count"].sum())
                    row.append(val)
                # Total for this classification
                mask_all = grp[status_col] == lbl
                row.append(int(grp.loc[mask_all, "count"].sum()))
            detail_data.append(row)

        # Sort by total descending
        detail_data.sort(key=lambda x: x[1], reverse=True)
        for row in detail_data:
            rows_out.append(row)

        # Grand Total row
        grand = ["Grand Total", len(df)]
        for lbl in detail_labels:
            for age_cat in age_cats:
                mask = (grouped["Kategori_Umur"] == age_cat) & (
                    grouped[status_col] == lbl
                )
                grand.append(int(grouped.loc[mask, "count"].sum()))
            mask_all = grouped[status_col] == lbl
            grand.append(int(grouped.loc[mask_all, "count"].sum()))
        rows_out.append(grand)
        rows_out.append([])

        # --- SECTION 2: Detailed Percentage ---
        rows_out.append([f"Count of {status_col}", "Column Labels"])
        rows_out.append(sub_hdr1_1)
        sub_pct = ["Row Labels", ""]
        for lbl in detail_labels:
            sub_pct.extend(["Bawah 2 Tahun", "Bawah 5 Tahun", ""])
        rows_out.append(sub_pct)

        for row in detail_data:
            geo_name = row[0]
            total = row[1]
            pct_row = [geo_name, ""]
            idx = 2
            for lbl in detail_labels:
                for _ in age_cats:
                    pct_row.append(f"{row[idx] / total * 100:.2f}%" if total else "0%")
                    idx += 1
                pct_row.append(f"{row[idx] / total * 100:.2f}%" if total else "0%")
                idx += 1
            rows_out.append(pct_row)

        # Grand total pct
        g_total = grand[1]
        gpct = ["Grand Total", ""]
        idx = 2
        for lbl in detail_labels:
            for _ in age_cats:
                gpct.append(f"{grand[idx] / g_total * 100:.2f}%" if g_total else "0%")
                idx += 1
            gpct.append(f"{grand[idx] / g_total * 100:.2f}%" if g_total else "0%")
            idx += 1
        rows_out.append(gpct)
        rows_out.append([])

        # --- SECTION 3: Combined Count ---
        rows_out.append(
            [f"Count of {status_col.replace('_Status', '_COMBINE')}", "Column Labels"]
        )
        sub_c1 = ["", ""]
        for clbl in combine_order:
            sub_c1.extend([clbl, "", f"{clbl} Total"])
        sub_c1.append("Grand Total")
        rows_out.append(sub_c1)

        sub_c2 = ["Row Labels", ""]
        for clbl in combine_order:
            sub_c2.extend(["Bawah 2 Tahun", "Bawah 5 Tahun", ""])
        sub_c2.append("")
        rows_out.append(sub_c2)

        combine_data = []
        for row in detail_data:
            geo_name = row[0]
            total = row[1]
            crow = [geo_name, ""]
            for clbl in combine_order:
                codes = combine_map[clbl]
                for age_cat in age_cats:
                    s = 0
                    idx_base = 2
                    for i, lbl in enumerate(detail_labels):
                        if lbl in codes:
                            age_idx = age_cats.index(age_cat)
                            s += row[idx_base + i * (len(age_cats) + 1) + age_idx]
                    crow.append(s)
                # Total
                s_total = 0
                idx_base = 2
                for i, lbl in enumerate(detail_labels):
                    if lbl in codes:
                        s_total += row[
                            idx_base + i * (len(age_cats) + 1) + len(age_cats)
                        ]
                crow.append(s_total)
            crow.append(total)
            combine_data.append(crow)

        for crow in combine_data:
            rows_out.append(crow)

        # Grand total combined
        gcrow = ["Grand Total", ""]
        g_total = grand[1]
        for clbl in combine_order:
            codes = combine_map[clbl]
            for age_cat in age_cats:
                s = 0
                idx_base = 2
                for i, lbl in enumerate(detail_labels):
                    if lbl in codes:
                        age_idx = age_cats.index(age_cat)
                        s += grand[idx_base + i * (len(age_cats) + 1) + age_idx]
                gcrow.append(s)
            s_total = 0
            idx_base = 2
            for i, lbl in enumerate(detail_labels):
                if lbl in codes:
                    s_total += grand[idx_base + i * (len(age_cats) + 1) + len(age_cats)]
            gcrow.append(s_total)
        gcrow.append(g_total)
        rows_out.append(gcrow)
        rows_out.append([])

        # --- SECTION 4: Combined Percentage ---
        rows_out.append(
            [f"Count of {status_col.replace('_Status', '_COMBINE')}", "Column Labels"]
        )
        rows_out.append(sub_c1)
        rows_out.append(sub_c2)

        for crow in combine_data:
            geo_name = crow[0]
            total = crow[-1]
            cpct = [geo_name, ""]
            idx = 2
            for clbl in combine_order:
                for _ in age_cats:
                    cpct.append(f"{crow[idx] / total * 100:.2f}%" if total else "0%")
                    idx += 1
                cpct.append(f"{crow[idx] / total * 100:.2f}%" if total else "0%")
                idx += 1
            cpct.append("100.00%")
            rows_out.append(cpct)

        gcpct = ["Grand Total", ""]
        idx = 2
        for clbl in combine_order:
            for _ in age_cats:
                gcpct.append(f"{gcrow[idx] / g_total * 100:.2f}%" if g_total else "0%")
                idx += 1
            gcpct.append(f"{gcrow[idx] / g_total * 100:.2f}%" if g_total else "0%")
            idx += 1
        gcpct.append("100.00%")
        rows_out.append(gcpct)

        return rows_out

    # ── WAZ config ─────────────────────────────────────────────────────────────
    waz_detail = [
        "kurang_berat_badan_teruk",
        "kurang_berat_badan",
        "risiko_kurang_berat_badan",
        "berat_badan_normal",
        "mungkin_masalah_pertumbuhan",
    ]
    waz_combine = {
        "Kurang Berat Badan": ["kurang_berat_badan_teruk", "kurang_berat_badan"],
        "risiko_kurang_berat_badan": ["risiko_kurang_berat_badan"],
        "berat_badan_normal": ["berat_badan_normal"],
        "mungkin_masalah_pertumbuhan": ["mungkin_masalah_pertumbuhan"],
    }
    waz_combine_order = [
        "Kurang Berat Badan",
        "risiko_kurang_berat_badan",
        "berat_badan_normal",
        "mungkin_masalah_pertumbuhan",
    ]

    # ── HAZ config ─────────────────────────────────────────────────────────────
    haz_detail = [
        "bantut_teruk",
        "bantut",
        "risiko_bantut",
        "normal",
        "mungkin_masalah_endokrin",
    ]
    haz_combine = {
        "Bantut": ["bantut_teruk", "bantut"],
        "risiko_bantut": ["risiko_bantut"],
        "normal": ["normal"],
        "mungkin_masalah_endokrin": ["mungkin_masalah_endokrin"],
    }
    haz_combine_order = [
        "Bantut",
        "risiko_bantut",
        "normal",
        "mungkin_masalah_endokrin",
    ]

    # ── BAZ config ─────────────────────────────────────────────────────────────
    baz_detail = [
        "susut_teruk",
        "susut",
        "berisiko_susut",
        "normal",
        "risiko_lebih_berat_badan",
        "berlebihan_berat_badan",
        "obes",
    ]
    baz_combine = {
        "Susut": ["susut_teruk", "susut"],
        "berisiko_susut": ["berisiko_susut"],
        "normal": ["normal"],
        "risiko_lebih_berat_badan": ["risiko_lebih_berat_badan"],
        "Berlebihan Berat Badan": ["berlebihan_berat_badan", "obes"],
    }
    baz_combine_order = [
        "Susut",
        "berisiko_susut",
        "normal",
        "risiko_lebih_berat_badan",
        "Berlebihan Berat Badan",
    ]

    # Build pivot tabs
    waz_negeri = _build_pivot_tab(
        state_col,
        "Negeri",
        "WAZ_Status",
        waz_detail,
        waz_combine,
        waz_combine_order,
        "WAZ Negeri",
    )
    waz_daerah = _build_pivot_tab(
        district_col,
        "Daerah",
        "WAZ_Status",
        waz_detail,
        waz_combine,
        waz_combine_order,
        "WAZ Daerah",
    )
    haz_negeri = _build_pivot_tab(
        state_col,
        "Negeri",
        "HAZ_Status",
        haz_detail,
        haz_combine,
        haz_combine_order,
        "HAZ Negeri",
    )
    haz_daerah = _build_pivot_tab(
        district_col,
        "Daerah",
        "HAZ_Status",
        haz_detail,
        haz_combine,
        haz_combine_order,
        "HAZ Daerah",
    )
    baz_negeri = _build_pivot_tab(
        state_col,
        "Negeri",
        "BAZ_Status",
        baz_detail,
        baz_combine,
        baz_combine_order,
        "BAZ Negeri",
    )
    baz_daerah = _build_pivot_tab(
        district_col,
        "Daerah",
        "BAZ_Status",
        baz_detail,
        baz_combine,
        baz_combine_order,
        "BAZ Daerah",
    )

    # Write workbook
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(exec_rows).to_excel(
            writer, sheet_name="Executive Summary", index=False, header=False
        )
        _style_sheet(writer.sheets["Executive Summary"], header_row=4, freeze_row=5)

        pd.DataFrame(rules_detail).to_excel(
            writer, sheet_name="Cleaning Rules", index=False, header=False
        )
        _style_sheet(writer.sheets["Cleaning Rules"], header_row=3, freeze_row=4)

        pd.DataFrame(dropped_rows).to_excel(
            writer, sheet_name="Records Dropped", index=False, header=False
        )
        _style_sheet(writer.sheets["Records Dropped"], header_row=3, freeze_row=4)

        # 6 pivot tabs
        pivot_tabs = [
            ("WAZ Negeri", waz_negeri),
            ("WAZ Daerah", waz_daerah),
            ("HAZ Negeri", haz_negeri),
            ("HAZ Daerah", haz_daerah),
            ("BAZ Negeri", baz_negeri),
            ("BAZ Daerah", baz_daerah),
        ]
        # Pivot tabs each contain four section blocks (Detailed Count,
        # Detailed Percentage, Combined Count, Combined Percentage). Each
        # block has three header rows: "Count of …" (section title),
        # classification labels, and "Row Labels" (age categories). Previously
        # only the "Row Labels" row got blue-filled and no borders / column
        # widths / freeze panes were applied, so sections 2-4 looked unstyled
        # compared to the Executive Summary tab. This pass applies the same
        # visual treatment everywhere, plus the new KKM Navy palette.
        navy_hex = "1B2A4A"  # KKM Navy — header fill
        navy_text_hex = "0F1B2F"  # KKM Navy Dark — bold title text
        gold_text_hex = "C8962E"  # KKM Gold  — section-title accent
        thin = Side(style="thin", color="D8DFEC")
        all_borders = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(
            start_color=navy_hex, end_color=navy_hex, fill_type="solid"
        )
        section_font = Font(bold=True, size=12, color=gold_text_hex)
        title_font = Font(bold=True, size=14, color=navy_text_hex)

        def _style_header_row(row):
            for cell in row:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = all_borders

        for sheet_name, rows_data in pivot_tabs:
            pd.DataFrame(rows_data).to_excel(
                writer, sheet_name=sheet_name, index=False, header=False
            )
            ws = writer.sheets[sheet_name]

            # Pass 1: identify header rows so we can style each block's
            # three-row header consistently (section title + sub-header +
            # Row Labels row).
            header_rows: set[int] = set()
            for i, row in enumerate(
                ws.iter_rows(min_row=1, max_row=ws.max_row), start=1
            ):
                val = str(row[0].value or "")
                if val.startswith("Count of"):
                    # mark the next two rows as part of this block's header
                    header_rows.update({i, i + 1, i + 2})
                if val == "Row Labels":
                    header_rows.add(i)

            # Pass 2: apply styles
            for i, row in enumerate(
                ws.iter_rows(min_row=1, max_row=ws.max_row), start=1
            ):
                val = str(row[0].value or "")
                if val.startswith("Count of"):
                    # Section title row — gold-accented bold, no fill so it
                    # reads as a banner above the navy-filled column header.
                    for cell in row:
                        cell.font = section_font
                        cell.border = all_borders
                elif i in header_rows:
                    _style_header_row(row)
                elif val == "Grand Total":
                    for cell in row:
                        cell.font = Font(bold=True, color=navy_text_hex)
                        cell.border = all_borders
                else:
                    for cell in row:
                        cell.border = all_borders

            # Column widths sized to longest value (capped at 45).
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[letter]:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[letter].width = min(max_len + 4, 45)

            # Freeze panes below the first block's three header rows so the
            # column categories stay visible while scrolling.
            ws.freeze_panes = ws.cell(row=4, column=2)

        # Title-style the first cell of every sheet (Executive Summary, etc.).
        for ws in writer.sheets.values():
            if ws.cell(1, 1).value:
                ws.cell(1, 1).font = title_font

    output.seek(0)
    return output


@app.get("/clean/download-report/{cache_id}")
async def download_report_endpoint(
    cache_id: str,
    data_type: str = Query("myvass"),
):
    """Download Data Quality Report (5-tab Excel) from cached cleaning results."""
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(404, "Cached data not found \u2014 please re-run cleaning.")

    # Flag-then-filter: the report's analytical sheets (WAZ/HAZ/BAZ pivots,
    # executive summary) must describe the ANALYZABLE population so they match
    # the dashboard and the default download. The "Records Dropped" tab reads
    # its per-rule counts from `stats`, so it is unaffected by this projection.
    cleaned_df = _analysis_view(entry["df"])
    stats = entry["stats"]
    timestamp = pd.Timestamp.now().strftime("%Y%m%d")
    report_buf = _build_quality_report(cleaned_df, stats, data_type)

    return StreamingResponse(
        iter([report_buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{data_type.upper()}_Data_Quality_Report_{timestamp}.xlsx"'
        },
    )


# ─── AI ENDPOINTS ─────────────────────────────────────────────────────────────


class ReportRequest(BaseModel):
    cache_id: str
    eda_result: dict
    narrative: dict = {}
    kpi_result: dict | None = None


class _AIBody(BaseModel):
    """Optional JSON body for AI endpoints (frontend posts { question })."""

    question: str = ""


def _localize(v, lang: str = "en") -> str:
    """Collapse a possibly-bilingual value ({bm,en}) into a plain string."""
    if isinstance(v, dict):
        return str(v.get(lang) or v.get("en") or v.get("bm") or "").strip()
    if v is None:
        return ""
    return str(v)


def _format_narrative(n: dict) -> str:
    """Render the structured narrative dict into readable text for the chat UI."""
    if not isinstance(n, dict):
        return str(n)
    parts: list[str] = []
    summary = _localize(n.get("executive_summary"))
    if summary:
        parts.append(summary)
    insights = n.get("insights_5w1h") or {}
    if isinstance(insights, dict):
        labels = {
            "who": "Who",
            "what": "What",
            "when": "When",
            "where": "Where",
            "why": "Why",
            "how": "How",
        }
        lines = [
            f"• {labels.get(k, k.title())}: {_localize(v)}"
            for k, v in insights.items()
            if _localize(v)
        ]
        if lines:
            parts.append("Insights (5W1H):\n" + "\n".join(lines))
    recs = n.get("recommendations") or []
    if isinstance(recs, list) and recs:
        rec_lines = []
        for r in recs:
            if not isinstance(r, dict):
                continue
            action = r.get("action") or ""
            prio = r.get("priority") or ""
            detail = _localize(r)
            tag = f"[{prio}] " if prio else ""
            rec_lines.append(f"• {tag}{action}: {detail}".strip())
        if rec_lines:
            parts.append("Recommendations:\n" + "\n".join(rec_lines))
    return "\n\n".join(p for p in parts if p) or "No narrative produced."


@app.post("/ai/narrative")
def ai_narrative(
    cache_id: str = Query(...),
    chat_id: str | None = Query(
        None, description="Optional chat session to persist this exchange into."
    ),
):
    """Generate an AI narrative for the cleaned dataset referenced by cache_id.

    When chat_id is provided the AI message is also appended to that chat
    session so the user's transcript survives page reloads. Omit chat_id
    for the legacy stateless behaviour.
    """
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(404, "cache_id not found — re-upload or re-run cleaning.")
    df = entry["df"]
    source_type = (entry.get("stats") or {}).get("source_type") or "myvass"
    try:
        eda_result = run_eda_auto(df, source_type)
        narrative = generate_narrative(eda_result)
    except OllamaError as e:
        # AI model unreachable / still loading — give the user an actionable
        # message (and a 503 so the frontend can offer a retry) rather than a
        # bare 500. Common after a long idle while the model reloads.
        logger.warning("AI narrative unavailable for %s: %s", cache_id, e)
        raise HTTPException(
            503,
            "AI model is starting up or unavailable — please wait a few seconds "
            "and try again. (Model warming after idle.)",
        )
    except Exception as e:
        raise HTTPException(500, f"Narrative generation failed: {e}")
    # Persist by cache_id (overwrite — regeneration is latest-wins) so the
    # report endpoints embed the same insights the AI page shows.
    _cache_set_narrative(cache_id, narrative)
    formatted = _format_narrative(narrative)

    if chat_id:
        _chat_append_message(
            chat_id, role="narrative", content=formatted, data_json=narrative
        )

    return {"narrative": formatted, "raw": narrative}


@app.post("/ai/nlq")
def ai_nlq(
    cache_id: str = Query(...),
    chat_id: str | None = Query(
        None, description="Optional chat session to persist this exchange into."
    ),
    body: _AIBody | None = None,
):
    """Answer a natural-language question against the cleaned dataset.

    When chat_id is provided both the user question and the AI answer are
    appended to that chat session. Omit chat_id for legacy stateless use.
    """
    entry = _cache_get(cache_id)
    if entry is None:
        raise HTTPException(404, "cache_id not found — re-upload or re-run cleaning.")
    df = entry["df"]
    question = (body.question if body else "") or ""
    try:
        result = answer_query(question, df)
    except Exception as e:
        raise HTTPException(500, f"NLQ failed: {e}")
    answer = _localize(result.get("answer"))

    if chat_id and question:
        _chat_append_message(chat_id, role="user", content=question)
        _chat_append_message(
            chat_id,
            role="ai",
            content=answer or "",
            data_json={
                "data": result.get("result"),
                "chart_b64": result.get("chart_b64"),
            },
        )

    return {
        "answer": answer,
        "data": result.get("result"),
        "chart_b64": result.get("chart_b64"),
    }


# ── Chat sessions ────────────────────────────────────────────────────────────
# Anchored to a dataset (cache_id == Dataset.id). The /ai/nlq and
# /ai/narrative hooks above call _chat_append_message() when a chat_id is
# supplied so the user's transcript survives reloads.


class _ChatTitleBody(BaseModel):
    title: str


class _ChatMessageBody(BaseModel):
    role: str
    content: str
    data_json: dict | None = None


def _chat_append_message(
    chat_session_id: str,
    role: str,
    content: str,
    data_json: dict | None = None,
) -> None:
    """Best-effort append. Auto-titles the chat from the first user
    question when the title is still the placeholder "New chat" so the
    sidebar entry becomes meaningful without an explicit rename."""
    from .db.init_db import SessionLocal
    from .db.models import ChatSession as _ChatSession, ChatMessage as _ChatMessage

    if SessionLocal is None:
        return
    try:
        db = SessionLocal()
        try:
            cs = db.get(_ChatSession, chat_session_id)
            if cs is None:
                logger.warning(
                    "chat append: session %s not found; dropping message",
                    chat_session_id,
                )
                return
            db.add(
                _ChatMessage(
                    chat_session_id=chat_session_id,
                    role=role,
                    content=content,
                    data_json=data_json,
                )
            )
            cs.updated_at = datetime.utcnow()
            # Auto-title from the first user question, OR from the first AI
            # insight when the chat is narrative-only (so it stops reading as a
            # stack of identical "New chat" entries in the sidebar).
            if cs.title == "New chat" and role in ("user", "narrative"):
                first = next(
                    (ln for ln in (content or "").splitlines() if ln.strip()), ""
                )
                # Strip leading markdown heading / bullet markers.
                trimmed = first.strip().lstrip("#*->•").strip()
                if role == "narrative":
                    trimmed = f"AI Insight — {trimmed}" if trimmed else "AI Insight"
                cs.title = (
                    (trimmed[:60] + "…")
                    if len(trimmed) > 60
                    else (trimmed or "New chat")
                )
            db.commit()
        finally:
            db.close()
    except Exception as exc:  # pragma: no cover — defensive
        logger.warning("chat append failed for session %s: %s", chat_session_id, exc)


@app.get("/chats")
def list_chats(dataset_id: str = Query(...), db=Depends(get_db)):
    """List chat sessions for a dataset, newest activity first."""
    from .db.models import ChatSession as _ChatSession, ChatMessage as _ChatMessage
    from sqlalchemy import func

    rows = (
        db.query(
            _ChatSession,
            func.count(_ChatMessage.id).label("message_count"),
        )
        .outerjoin(_ChatMessage, _ChatMessage.chat_session_id == _ChatSession.id)
        .filter(_ChatSession.dataset_id == dataset_id)
        .group_by(_ChatSession.id)
        .order_by(_ChatSession.updated_at.desc())
        .all()
    )
    return [
        {
            "id": cs.id,
            "title": cs.title,
            "message_count": int(msg_count or 0),
            "created_at": cs.created_at.isoformat(),
            "updated_at": cs.updated_at.isoformat(),
        }
        for cs, msg_count in rows
    ]


@app.post("/chats")
def create_chat(dataset_id: str = Query(...), db=Depends(get_db)):
    """Create a new empty chat session for a dataset."""
    from .db.models import ChatSession as _ChatSession, Dataset as _Dataset

    if db.get(_Dataset, dataset_id) is None:
        raise HTTPException(404, f"dataset_id {dataset_id} not found")
    cs = _ChatSession(id=str(_uuid.uuid4()), dataset_id=dataset_id, title="New chat")
    db.add(cs)
    db.commit()
    _log_audit(action="chat.create", dataset_id=dataset_id, detail=f"chat_id={cs.id}")
    return {
        "id": cs.id,
        "title": cs.title,
        "created_at": cs.created_at.isoformat(),
        "updated_at": cs.updated_at.isoformat(),
    }


@app.get("/chats/{chat_id}")
def get_chat(chat_id: str, db=Depends(get_db)):
    """Return chat metadata + ordered messages."""
    from .db.models import ChatSession as _ChatSession

    cs = db.get(_ChatSession, chat_id)
    if cs is None:
        raise HTTPException(404, f"chat_id {chat_id} not found")
    return {
        "id": cs.id,
        "dataset_id": cs.dataset_id,
        "title": cs.title,
        "created_at": cs.created_at.isoformat(),
        "updated_at": cs.updated_at.isoformat(),
        "messages": [
            {
                "id": m.id,
                "role": m.role,
                "content": m.content,
                "data_json": m.data_json,
                "created_at": m.created_at.isoformat(),
            }
            for m in cs.messages
        ],
    }


@app.patch("/chats/{chat_id}")
def rename_chat(chat_id: str, body: _ChatTitleBody, db=Depends(get_db)):
    """Rename a chat session."""
    from .db.models import ChatSession as _ChatSession

    cs = db.get(_ChatSession, chat_id)
    if cs is None:
        raise HTTPException(404, f"chat_id {chat_id} not found")
    cs.title = (body.title or "").strip()[:200] or cs.title
    cs.updated_at = datetime.utcnow()
    db.commit()
    return {"id": cs.id, "title": cs.title}


@app.delete("/chats/{chat_id}")
def delete_chat(chat_id: str, db=Depends(get_db)):
    """Delete a chat session — messages cascade away with it."""
    from .db.models import ChatSession as _ChatSession

    cs = db.get(_ChatSession, chat_id)
    if cs is None:
        raise HTTPException(404, f"chat_id {chat_id} not found")
    dataset_id = cs.dataset_id
    db.delete(cs)
    db.commit()
    _log_audit(action="chat.delete", dataset_id=dataset_id, detail=f"chat_id={chat_id}")
    return {"deleted": chat_id}


@app.post("/chats/{chat_id}/messages")
def post_chat_message(chat_id: str, body: _ChatMessageBody, db=Depends(get_db)):
    """Append a message to a chat session (manual write — most messages
    go in via the /ai/nlq and /ai/narrative auto-persist hooks)."""
    from .db.models import ChatSession as _ChatSession, ChatMessage as _ChatMessage

    cs = db.get(_ChatSession, chat_id)
    if cs is None:
        raise HTTPException(404, f"chat_id {chat_id} not found")
    if body.role not in ("user", "ai", "narrative"):
        raise HTTPException(400, "role must be one of: user, ai, narrative")
    m = _ChatMessage(
        chat_session_id=chat_id,
        role=body.role,
        content=body.content,
        data_json=body.data_json,
    )
    db.add(m)
    cs.updated_at = datetime.utcnow()
    if body.role == "user" and cs.title == "New chat":
        trimmed = (
            (body.content or "").strip().splitlines()[0] if body.content.strip() else ""
        )
        cs.title = (
            (trimmed[:60] + "…") if len(trimmed) > 60 else (trimmed or "New chat")
        )
    db.commit()
    return {"id": m.id, "created_at": m.created_at.isoformat()}


# ── Multi-Dataset Comparison ─────────────────────────────────────────────────

from backend.eda.compare import compare_datasets


@app.get("/datasets")
async def list_datasets(owner: str | None = Depends(_identity)):
    """List datasets in the library, scoped to the current identity.

    When an identity (X-User) is present, only that owner's datasets — plus
    legacy un-owned rows created before per-person scoping — are returned, so
    each person sees their own history across devices. With no identity header
    the full library is returned (no regression for un-headered callers)."""
    from backend.db.models import Dataset
    from backend.db.init_db import SessionLocal

    with SessionLocal() as db:
        q = db.query(Dataset)
        if owner is not None:
            q = q.filter((Dataset.owner == owner) | (Dataset.owner.is_(None)))
        datasets = q.order_by(Dataset.created_at.desc()).all()
        return [
            {
                "id": ds.id,
                "name": ds.name,
                "filename": ds.filename,
                "source_type": ds.source_type,
                "row_count": ds.row_count,
                "quality_score": ds.quality_score,
                "created_at": ds.created_at.isoformat(),
            }
            for ds in datasets
        ]


class DatasetCompareRequest(BaseModel):
    dataset_ids: list[str]


class DatasetDeleteRequest(BaseModel):
    dataset_ids: list[str]


def _delete_datasets(db, dataset_ids: list[str]) -> dict:
    """Hard-delete datasets and their dependents in FK-safe order.
    Order: AnalysisResult -> Session -> Dataset, then evict the disk/memory
    cache. Commits internally before evicting cache; the caller is responsible
    for rolling back on exception."""
    from backend.db.models import (
        Dataset,
        Session as _Session,
        AnalysisResult,
        EntityLinkage,
    )

    deleted: list[str] = []
    not_found: list[str] = []
    for ds_id in dataset_ids:
        ds = db.get(Dataset, ds_id)
        if ds is None:
            not_found.append(ds_id)
            continue
        session_ids = [
            s.id for s in db.query(_Session).filter(_Session.dataset_id == ds_id).all()
        ]
        if session_ids:
            db.query(AnalysisResult).filter(
                AnalysisResult.session_id.in_(session_ids)
            ).delete(synchronize_session=False)
            db.query(_Session).filter(_Session.dataset_id == ds_id).delete(
                synchronize_session=False
            )
        db.query(EntityLinkage).filter(EntityLinkage.dataset_id == ds_id).update(
            {"dataset_id": None}, synchronize_session=False
        )
        db.delete(ds)
        deleted.append(ds_id)

    db.commit()
    for ds_id in deleted:
        _cache_evict(ds_id)
    return {"deleted": deleted, "not_found": not_found}


@app.post("/datasets/compare")
async def datasets_compare(req: DatasetCompareRequest, db=Depends(get_db)):
    """Compare 2+ datasets side-by-side. Returns quality and indicator deltas.

    Quality comes from the persisted Dataset row (always present). The four
    nutrition prevalence rates are recomputed on the fly from each dataset's
    cached cleaned frame — the same compute_kpi_dashboard path the dashboard
    uses — because per-dataset indicators are not persisted anywhere. A
    dataset whose cache has been evicted still contributes its quality score;
    its indicators degrade to empty (rendered as "—") rather than failing the
    whole comparison.

    Prior bug: this queried a non-existent AnalysisResult.dataset_id column
    (500) and read a result_json shape that is never written, so the modal
    always came back empty.
    """
    from backend.db.models import Dataset

    if len(req.dataset_ids) < 2:
        raise HTTPException(400, "Provide at least 2 dataset_ids.")

    targets = _load_kpi_targets(db)
    summaries = []
    for ds_id in req.dataset_ids:
        ds = db.get(Dataset, ds_id)
        if ds is None:
            continue
        indicators: dict[str, float] = {}
        entry = _cache_get(ds_id)
        if entry is not None and entry.get("df") is not None:
            try:
                kpi = compute_kpi_dashboard(
                    entry["df"], npan=targets["npan"], who=targets["who"]
                )
                # kpi indicator `key` is the flag (stunting/wasting/…) and
                # `actual` is a percentage; compare_datasets expects fractions
                # keyed by `<flag>_rate` (stunting_rate, wasting_rate, …).
                for i in kpi.get("indicators", []):
                    key, actual = i.get("key"), i.get("actual")
                    if key and actual is not None:
                        indicators[f"{key}_rate"] = float(actual) / 100.0
            except Exception as exc:  # pragma: no cover — defensive
                logger.warning(
                    "compare: indicator compute failed for %s: %s", ds_id, exc
                )
        summaries.append(
            {
                "dataset_id": ds_id,
                "name": ds.name,
                "source_type": normalize_schema_type(ds.source_type or "general"),
                "quality_score": _coerce_float(ds.quality_score),
                "indicators": indicators,
                "created_at": ds.created_at.isoformat() if ds.created_at else None,
            }
        )

    # Oldest → latest so compare_datasets' "latest vs earliest" deltas and the
    # OLS trend read chronologically regardless of selection order. ISO UTC
    # strings sort lexically == chronologically.
    summaries.sort(key=lambda s: s["created_at"] or "")

    return JSONResponse(content=json_safe(compare_datasets(summaries)))


@app.post("/datasets/delete")
async def datasets_delete(
    req: DatasetDeleteRequest, owner: str | None = Depends(_identity)
):
    """Permanently delete datasets: DB rows + sessions/analysis + cache."""
    from backend.db.init_db import SessionLocal

    if not req.dataset_ids:
        raise HTTPException(400, "Provide at least 1 dataset_id.")
    with SessionLocal() as db:
        try:
            result = _delete_datasets(db, req.dataset_ids)
        except Exception as e:
            db.rollback()
            raise HTTPException(500, f"Delete failed: {e}")
    _log_audit(action="dataset.delete", detail=f"ids={result['deleted']}", actor=owner)
    return result


# ── Entity Resolution ────────────────────────────────────────────────────────

from backend.ml.entity import (
    link_records,
    link_records_v2,
    persist_linkage,
    _normalise_ic,
)


class EntityLinkRequest(BaseModel):
    dataset_ids: list[str]


class EntityLinkV2Request(BaseModel):
    dataset_ids: list[str]
    fuzzy_ic: bool = True
    fuzzy_ic_max_distance: int = 1
    name_dob_boost: bool = True
    # New in v2.1 — close the gap to spec (probabilistic name + DOB +
    # location signals + contradiction surfacing). All default-on; toggle
    # off to recover v2-original behaviour.
    name_fuzzy: bool = True
    name_fuzzy_threshold: float = 0.85
    dob_tolerance_days: int = 1
    location_boost: bool = True
    min_confidence: float = 0.0  # 0.0 = include unmatched singles
    max_groups: int = 500  # cap response size — UI paginates


class EntityRecordsSyncRequest(BaseModel):
    dataset_ids: list[str]


_IC_COL_CANDIDATES = (
    "IC_NO_PASSPORT",
    "IC",
    "NRIC",
    "MyKID",
    "ic_no",
    "no_kp",
    "no_ic",
)
_NAME_COL_CANDIDATES = ("NAMA", "name", "NAMA_PESERTA", "FULL_NAME", "nama_kanak_kanak")
_DOB_COL_CANDIDATES = ("Tarikh_Lahir", "DOB", "TARIKH_LAHIR", "dob", "date_of_birth")
_GENDER_COL_CANDIDATES = ("jantina", "JANTINA", "gender", "GENDER", "sex")
_STATE_COL_CANDIDATES = ("negeri", "NEGERI", "state", "STATE")
_DISTRICT_COL_CANDIDATES = ("daerah", "DAERAH", "district", "DISTRICT", "kawasan")
_MEASURE_DATE_CANDIDATES = (
    "Tarikh_Pengukuran",
    "TARIKH_PENGUKURAN",
    "tarikh_ukur_dt",
    "tarikh_ukur",
    "measure_date",
)
# Anthropometrics + z-scores — coerced via pd.to_numeric for the timeline.
_NUMERIC_TIMELINE_COLS = ("berat_kg", "tinggi_cm", "bmi", "waz", "haz", "baz")


def _pick_col(df_cols: list[str], candidates: tuple[str, ...]) -> str | None:
    """Return the first matching column name (case-insensitive) or None."""
    lower = {c.lower(): c for c in df_cols}
    for cand in candidates:
        hit = lower.get(cand.lower())
        if hit:
            return hit
    return None


def _records_from_cached(
    cache_id: str,
    dataset_id: str,
    source_type: str,
    dataset_created_at: "datetime | None" = None,
) -> list[dict]:
    """Pull child-level records from a cached DataFrame so v2 linkage can
    do name / DOB / location / contradiction work without re-querying the
    DB. Carries through IC + name + DOB + gender + state + district +
    measurement date + anthropometrics + z-scores. None survives end-to-end
    — no `astype(str)` → 'nan' false-conflict surface."""
    entry = _cache_get(cache_id)
    if entry is None:
        return []
    df = entry["df"]
    if df is None or df.empty:
        return []
    cols = list(df.columns)
    ic_c = _pick_col(cols, _IC_COL_CANDIDATES)
    name_c = _pick_col(cols, _NAME_COL_CANDIDATES)
    dob_c = _pick_col(cols, _DOB_COL_CANDIDATES)
    gender_c = _pick_col(cols, _GENDER_COL_CANDIDATES)
    state_c = _pick_col(cols, _STATE_COL_CANDIDATES)
    district_c = _pick_col(cols, _DISTRICT_COL_CANDIDATES)
    measure_c = _pick_col(cols, _MEASURE_DATE_CANDIDATES)
    has_year = "tahun_ukur" in cols
    has_month = "bulan_ukur" in cols
    if ic_c is None:
        return []

    def _str_or_none(v) -> str | None:
        """NaN/None/'nan'/'none' → None; otherwise stripped str."""
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        s = str(v).strip()
        if not s or s.lower() in ("nan", "none", "nat"):
            return None
        return s

    def _measure_date_for(idx: int) -> str | None:
        """Prefer the explicit column; fall back to tahun_ukur+bulan_ukur."""
        if measure_c is not None:
            v = df.iloc[idx][measure_c]
            s = _str_or_none(v)
            if s:
                return s
        if has_year and has_month:
            y = df.iloc[idx]["tahun_ukur"]
            m = df.iloc[idx]["bulan_ukur"]
            try:
                yi, mi = int(float(y)), int(float(m))
                if 1900 <= yi <= 2100 and 1 <= mi <= 12:
                    return f"{yi:04d}-{mi:02d}-01"
            except (TypeError, ValueError):
                return None
        return None

    out: list[dict] = []
    n = len(df)
    for i in range(n):
        rec: dict = {
            "ic": _str_or_none(df.iloc[i][ic_c]) or "",
            "source_type": normalize_schema_type(source_type or "general"),
            "dataset_id": dataset_id,
            "name": _str_or_none(df.iloc[i][name_c]) if name_c else None,
            "dob": _str_or_none(df.iloc[i][dob_c]) if dob_c else None,
            "gender": _str_or_none(df.iloc[i][gender_c]) if gender_c else None,
            "state": _str_or_none(df.iloc[i][state_c]) if state_c else None,
            "district": _str_or_none(df.iloc[i][district_c]) if district_c else None,
            "measure_date": _measure_date_for(i),
        }
        # Numerics — coerce per-cell so NaN becomes None rather than 'nan'.
        for c in _NUMERIC_TIMELINE_COLS:
            if c not in cols:
                rec[c] = None
                continue
            v = df.iloc[i][c]
            if v is None or (isinstance(v, float) and pd.isna(v)):
                rec[c] = None
                continue
            try:
                rec[c] = float(v)
            except (TypeError, ValueError):
                rec[c] = None
        out.append(rec)
    # Numeric keys we surface to the timeline use UI-friendlier names.
    for r in out:
        r["weight_kg"] = r.pop("berat_kg", None)
        r["height_cm"] = r.pop("tinggi_cm", None)
        # bmi/waz/haz/baz keep their names.
        if dataset_created_at is not None:
            r["_dataset_created_at"] = dataset_created_at
    return out


@app.post("/entity/records/sync")
async def entity_records_sync(req: EntityRecordsSyncRequest, db=Depends(get_db)):
    """P2-2: Backfill child_record table from currently cached datasets."""
    from backend.db.models import Dataset

    result = {}
    for ds_id in req.dataset_ids:
        entry = _cache_get(ds_id)
        if entry is None:
            result[ds_id] = 0
            continue
        try:
            recs = _records_from_cached(
                ds_id, ds_id, entry["stats"].get("source_type", "general")
            )
            count = _persist_child_records(
                db, ds_id, entry["stats"].get("source_type", "general"), recs
            )
            result[ds_id] = count
        except Exception as exc:
            logger.warning("Sync failed for dataset %s: %s", ds_id, exc)
            result[ds_id] = 0
    return {"persisted_counts": result}


@app.post("/entity/link/all")
async def entity_link_all(req: EntityLinkV2Request, db=Depends(get_db)):
    """P2-3: Link across ALL persisted datasets (or filtered subset).

    Uses the durable child_record store instead of volatile cache.
    Same response shape as /entity/link/v2.
    """
    from backend.db.init_db import SessionLocal

    # Read from durable store
    with SessionLocal() as db_session:
        records, dataset_created_at_by_id = _records_from_store(
            db_session, req.dataset_ids if req.dataset_ids else None
        )

    if not records:
        return {
            "total_groups": 0,
            "linked_groups": 0,
            "unlinked": 0,
            "datasets": [],
            "profiles": [],
        }

    # Run v2 linkage
    groups = link_records_v2(
        records,
        fuzzy_ic=req.fuzzy_ic,
        fuzzy_ic_max_distance=req.fuzzy_ic_max_distance,
        name_dob_boost=req.name_dob_boost,
        name_fuzzy=req.name_fuzzy,
        name_fuzzy_threshold=req.name_fuzzy_threshold,
        dob_tolerance_days=req.dob_tolerance_days,
        location_boost=req.location_boost,
        min_confidence=req.min_confidence,
        dataset_created_at_by_id=dataset_created_at_by_id,
    )

    # Build response matching /entity/link/v2 shape
    datasets_map = {}
    for rec in records:
        ds_id = rec["dataset_id"]
        if ds_id not in datasets_map:
            datasets_map[ds_id] = {
                "dataset_id": ds_id,
                "filename": ds_id,  # Fallback; could enrich from Dataset table
                "source_type": rec["source_type"],
                "records": 0,
            }
        datasets_map[ds_id]["records"] += 1

    profiles = groups[: req.max_groups] if req.max_groups else groups

    # P2-5: Add reconciliation summary
    reconciliation = _reconciliation_summary(groups)

    # P2-4: Persist linkage run for audit (best-effort)
    run_id = None
    try:
        run_id = _persist_linkage_run(
            db,
            groups,
            params={
                "fuzzy_ic": req.fuzzy_ic,
                "fuzzy_ic_max_distance": req.fuzzy_ic_max_distance,
                "name_dob_boost": req.name_dob_boost,
                "name_fuzzy": req.name_fuzzy,
                "name_fuzzy_threshold": req.name_fuzzy_threshold,
                "dob_tolerance_days": req.dob_tolerance_days,
                "location_boost": req.location_boost,
                "min_confidence": req.min_confidence,
            },
            dataset_ids=req.dataset_ids if req.dataset_ids else [],
        )
    except Exception as exc:
        logger.warning("Linkage run persistence failed: %s", exc)

    _log_audit(
        action="entity.link.all",
        detail=f"linked {len(profiles)} groups"
        + (f" (run_id={run_id})" if run_id else ""),
    )

    return {
        "total_groups": len(groups),
        "linked_groups": sum(1 for g in groups if len(g["sources"]) > 1),
        "unlinked": sum(1 for g in groups if len(g["sources"]) == 1),
        "datasets": list(datasets_map.values()),
        "profiles": profiles,
        "reconciliation": reconciliation,
        "run_id": run_id,
    }


@app.get("/entity/link/all/worklist")
async def entity_link_all_worklist(
    worklist_type: str = Query("conflicts", description="conflicts|duplicates"),
    db=Depends(get_db),
):
    """P2-5: Export worklist CSV for conflicts or duplicates."""
    from backend.db.init_db import SessionLocal

    # Re-run linkage to get groups (in production, would cache or use persisted runs)
    with SessionLocal() as db_session:
        records, _ = _records_from_store(db_session, None)

    if not records:
        return Response(content="", media_type="text/csv")

    groups = link_records_v2(records, min_confidence=0.0)

    buffer = io.StringIO()
    buffer.write(
        "group_index,ic,name,dob,source_type,dataset_id,confidence,field,severity,values\n"
    )

    if worklist_type == "conflicts":
        for gi, g in enumerate(groups):
            if not g.get("conflicts"):
                continue
            for conflict in g["conflicts"]:
                values_str = "|".join(
                    f"{v['source_type']}:{v['value']}" for v in conflict["values"]
                )
                for src in g["sources"]:
                    buffer.write(
                        f"{gi},{src.get('ic', '')},{src.get('name', '')},{src.get('dob', '')},"
                        f"{src['source_type']},{src['dataset_id']},{g['confidence']},"
                        f'{conflict["field"]},{conflict["severity"]},"{values_str}"\n'
                    )
    elif worklist_type == "duplicates":
        for gi, g in enumerate(groups):
            if len(g["sources"]) <= 1:
                continue
            sources = g["sources"]
            for src in sources:
                buffer.write(
                    f"{gi},{src.get('ic', '')},{src.get('name', '')},{src.get('dob', '')},"
                    f"{src['source_type']},{src['dataset_id']},{g['confidence']},,,\n"
                )

    buffer.seek(0)
    filename = f"SmartDQC_worklist_{worklist_type}.csv"
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/entity/link/runs")
async def list_linkage_runs(db=Depends(get_db)):
    """P2-4: List all linkage runs."""
    from backend.db.models import LinkageRun

    runs = db.query(LinkageRun).order_by(LinkageRun.created_at.desc()).limit(50).all()
    return {
        "runs": [
            {
                "id": r.id,
                "dataset_ids": r.dataset_ids,
                "total_groups": r.total_groups,
                "linked_groups": r.linked_groups,
                "created_at": r.created_at.isoformat(),
            }
            for r in runs
        ]
    }


@app.get("/entity/link/runs/{run_id}")
async def get_linkage_run(run_id: int, db=Depends(get_db)):
    """P2-4: Get details of a specific linkage run."""
    from backend.db.models import LinkageRun, LinkageMember

    run = db.get(LinkageRun, run_id)
    if not run:
        raise HTTPException(404, "Linkage run not found")

    members = db.query(LinkageMember).filter(LinkageMember.run_id == run_id).all()

    # Group members by group_index
    groups: dict[int, list] = {}
    for m in members:
        if m.group_index not in groups:
            groups[m.group_index] = []
        groups[m.group_index].append(
            {
                "ic_norm": m.ic_norm,
                "source_type": m.source_type,
                "dataset_id": m.dataset_id,
                "name": m.name,
                "dob": m.dob,
                "confidence": m.confidence,
                "match_reasons": m.match_reasons,
            }
        )

    return {
        "id": run.id,
        "params": run.params_json,
        "dataset_ids": run.dataset_ids,
        "total_groups": run.total_groups,
        "linked_groups": run.linked_groups,
        "created_at": run.created_at.isoformat(),
        "groups": [
            {"group_index": gi, "members": members_list}
            for gi, members_list in sorted(groups.items())
        ],
    }


@app.post("/entity/link")
async def entity_link(req: EntityLinkRequest):
    """Link child records across 2+ datasets by IC. Writes to entity_linkage table."""
    from backend.db.models import Dataset, AnalysisResult
    from backend.db.models import Session as DbSession
    from backend.db.init_db import SessionLocal

    if len(req.dataset_ids) < 2:
        raise HTTPException(400, "Provide at least 2 dataset_ids to link.")

    records = []
    with SessionLocal() as db:
        for ds_id in req.dataset_ids:
            ds = db.get(Dataset, ds_id)
            if ds is None:
                continue
            ar = (
                db.query(AnalysisResult)
                .join(AnalysisResult.session)
                .filter(DbSession.dataset_id == ds_id)
                .order_by(AnalysisResult.created_at.desc())
                .first()
            )
            if ar is None or not ar.result_json:
                continue
            summary = ar.result_json.get("summary", {})
            records.append(
                {
                    "ic": summary.get("ic", ""),
                    "source_type": normalize_schema_type(ds.source_type or "general"),
                    "dataset_id": ds_id,
                    "name": summary.get("name", ""),
                    "dob": summary.get("dob", ""),
                }
            )

        groups = link_records(records)
        rows_written = persist_linkage(groups, db)

    linked = [g for g in groups if len(g["sources"]) > 1]
    unlinked = [g for g in groups if len(g["sources"]) == 1]
    return {
        "total_groups": len(groups),
        "linked_groups": len(linked),
        "unlinked": len(unlinked),
        "rows_written": rows_written,
        "profiles": linked[:50],
    }


def _run_v2_linkage(req: EntityLinkV2Request) -> dict:
    """Shared core for /entity/link/v2 and /entity/link/v2/export.

    Reads actual child-level rows from each dataset's cached DataFrame
    (not just one summary row like v1), runs fuzzy IC + name/dob boost
    matching, and returns rich profiles with confidence + reason chips."""
    from backend.db.models import Dataset
    from backend.db.init_db import SessionLocal

    if len(req.dataset_ids) < 2:
        raise HTTPException(400, "Provide at least 2 dataset_ids to link.")

    all_records: list[dict] = []
    dataset_meta: list[dict] = []
    dataset_created_at_by_id: dict[str, datetime] = {}
    with SessionLocal() as db:
        for ds_id in req.dataset_ids:
            ds = db.get(Dataset, ds_id)
            if ds is None:
                continue
            recs = _records_from_cached(
                ds_id,
                ds_id,
                normalize_schema_type(ds.source_type or "general"),
                dataset_created_at=ds.created_at,
            )
            if ds.created_at is not None:
                dataset_created_at_by_id[ds_id] = ds.created_at
            dataset_meta.append(
                {
                    "dataset_id": ds_id,
                    "filename": ds.filename,
                    "source_type": ds.source_type,
                    "records": len(recs),
                    # ISO string so the UI can render "latest from X" labels.
                    "created_at": ds.created_at.isoformat() if ds.created_at else None,
                }
            )
            all_records.extend(recs)

    if not all_records:
        return {
            "total_groups": 0,
            "linked_groups": 0,
            "unlinked": 0,
            "datasets": dataset_meta,
            "profiles": [],
            "warning": "No matchable rows found in the cached datasets — re-run cleaning if the datasets were uploaded a long time ago and may have been evicted from the cache.",
        }

    groups = link_records_v2(
        all_records,
        fuzzy_ic=req.fuzzy_ic,
        fuzzy_ic_max_distance=req.fuzzy_ic_max_distance,
        name_dob_boost=req.name_dob_boost,
        name_fuzzy=req.name_fuzzy,
        name_fuzzy_threshold=req.name_fuzzy_threshold,
        dob_tolerance_days=req.dob_tolerance_days,
        location_boost=req.location_boost,
        min_confidence=req.min_confidence,
        dataset_created_at_by_id=dataset_created_at_by_id,
    )

    # Sort: linked-by-confidence-desc first, unlinked at the bottom.
    groups.sort(key=lambda g: (-len(g["sources"]), -g["confidence"]))

    linked = sum(1 for g in groups if len(g["sources"]) > 1)
    return {
        "total_groups": len(groups),
        "linked_groups": linked,
        "unlinked": len(groups) - linked,
        "datasets": dataset_meta,
        "profiles": groups[: req.max_groups],
    }


@app.post("/entity/link/v2")
async def entity_link_v2(req: EntityLinkV2Request):
    """Cross-dataset entity resolution v2 — fuzzy IC + name/dob boost.

    Returns matched child profiles with a 0-1 confidence score and
    explanation chips (exact_ic / fuzzy_ic±N / name+dob / unmatched).
    """
    result = _run_v2_linkage(req)
    _log_audit(
        action="entity.link.v2",
        detail=(
            f"{len(req.dataset_ids)} datasets, "
            f"{result['linked_groups']} matched, "
            f"name_fuzzy={req.name_fuzzy}({req.name_fuzzy_threshold}), "
            f"dob_tol={req.dob_tolerance_days}d, "
            f"loc_boost={req.location_boost}"
        ),
    )
    return result


@app.post("/entity/link/v2/export")
async def entity_link_v2_export(req: EntityLinkV2Request):
    """Return the v2 linkage result as a CSV download — one row per
    (group, source) pair so analysts can pivot on confidence and reasons."""
    result = _run_v2_linkage(req)
    import csv
    import io

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(
        [
            "group_ic",
            "canonical_name",
            "canonical_dob",
            "canonical_gender",
            "canonical_state",
            "canonical_district",
            "confidence",
            "match_reasons",
            "conflict_fields",
            "source_type",
            "dataset_id",
            "source_name",
            "source_dob",
            "source_ic",
            "source_gender",
            "source_state",
            "source_district",
        ]
    )
    for g in result["profiles"]:
        reasons = ";".join(g.get("match_reasons", []))
        canonical = (g.get("profile") or {}).get("canonical") or {}
        conflict_fields = ";".join(
            f"{c['field']}({c['severity']})" for c in g.get("conflicts", [])
        )
        for src in g["sources"]:
            w.writerow(
                [
                    g.get("ic", ""),
                    canonical.get("name") or "",
                    canonical.get("dob") or "",
                    canonical.get("gender") or "",
                    canonical.get("state") or "",
                    canonical.get("district") or "",
                    f"{g.get('confidence', 0.0):.2f}",
                    reasons,
                    conflict_fields,
                    src.get("source_type", ""),
                    src.get("dataset_id", ""),
                    src.get("name", ""),
                    src.get("dob", ""),
                    src.get("ic", ""),
                    src.get("gender", "") or "",
                    src.get("state", "") or "",
                    src.get("district", "") or "",
                ]
            )

    _log_audit(
        action="entity.link.v2.export",
        detail=f"{len(req.dataset_ids)} datasets, {result['linked_groups']} matched, csv",
    )
    return Response(
        content=buf.getvalue().encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="SmartDQC_Linkage.csv"'},
    )


@app.get("/dashboard/summary")
def dashboard_summary(owner: str | None = Depends(_identity), db=Depends(get_db)):
    """Aggregate stats for the dashboard landing, scoped to the current identity.

    Same owner contract as /datasets and /sessions: with an X-User header only
    that owner's datasets (plus legacy un-owned rows) are counted, so switching
    names yields that name's own latest_session/counts. No header = unscoped
    (no regression)."""
    from .db.models import Dataset

    # Exclude the placeholder dataset created by /ai/narrative — it has a
    # nil-UUID id and would otherwise surface as latest_session, causing the
    # frontend to request /kpi/dashboard?cache_id=0000... → 404.
    PLACEHOLDER_DATASET_ID = "00000000-0000-0000-0000-000000000000"
    q = db.query(Dataset)
    if owner is not None:
        q = q.filter((Dataset.owner == owner) | (Dataset.owner.is_(None)))
    datasets = [d for d in q.all() if d.id != PLACEHOLDER_DATASET_ID]
    if not datasets:
        return {
            "total_children": 0,
            "avg_quality_score": 0.0,
            "session_count": 0,
            "alerts": 0,
            "latest_session": None,
            "source_breakdown": {},
        }

    total_children = sum(d.row_count or 0 for d in datasets)
    scores = [d.quality_score for d in datasets if d.quality_score is not None]
    avg_quality = round(sum(scores) / len(scores), 1) if scores else 0.0
    alerts = sum(1 for d in datasets if (d.quality_score or 100) < 60)

    source_breakdown: dict[str, int] = {}
    for d in datasets:
        k = normalize_schema_type(d.source_type or "general")
        source_breakdown[k] = source_breakdown.get(k, 0) + 1

    latest = sorted(datasets, key=lambda d: d.created_at, reverse=True)[0]

    return {
        "total_children": total_children,
        "avg_quality_score": avg_quality,
        "session_count": len(datasets),
        "alerts": alerts,
        "latest_session": {
            "cache_id": latest.id,
            "filename": latest.filename,
            "source_type": latest.source_type,
            "created_at": latest.created_at.isoformat(),
        },
        "source_breakdown": source_breakdown,
    }


@app.get("/sessions")
def list_sessions(owner: str | None = Depends(_identity), db=Depends(get_db)):
    """List the 100 most recent cleaned sessions, scoped to the current identity.

    Legacy un-owned rows stay visible; with no identity header the list is
    unscoped (no regression)."""
    from .db.models import Dataset

    q = db.query(Dataset)
    if owner is not None:
        q = q.filter((Dataset.owner == owner) | (Dataset.owner.is_(None)))
    rows = q.order_by(Dataset.created_at.desc()).limit(100).all()
    return [
        {
            "cache_id": r.id,
            "name": r.name,
            "filename": r.filename,
            "source_type": r.source_type,
            "row_count": r.row_count or 0,
            "quality_score": r.quality_score or 0,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in rows
    ]


@app.get("/audit/log")
def get_audit_log(dataset_id: str | None = None, limit: int = 100, db=Depends(get_db)):
    """List audit entries with the user's username resolved via LEFT JOIN.

    dataset_id is a string FK to datasets.id (was incorrectly typed as int
    before, so the filter never matched anything)."""
    from .db.models import AuditLog, User

    q = (
        db.query(AuditLog, User.username)
        .outerjoin(User, AuditLog.user_id == User.id)
        .order_by(AuditLog.created_at.desc())
    )
    if dataset_id is not None:
        q = q.filter(AuditLog.dataset_id == dataset_id)
    rows = q.limit(limit).all()
    return [
        {
            "id": r.id,
            "action": r.action,
            "dataset_id": r.dataset_id,
            "detail": r.detail,
            "username": username,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r, username in rows
    ]


# ── Settings API ──────────────────────────────────────────────────────────────

import json as _json

_DEFAULT_THRESHOLDS = {
    "missing_rate_warn": 0.05,
    "missing_rate_fail": 0.15,
    "duplicate_rate_warn": 0.02,
    "duplicate_rate_fail": 0.10,
    "outlier_zscore_threshold": 3.0,
    # RAG band tolerances (C2). Stored as fractions above target. 0.20 / 0.30
    # preserve the historical target×1.20 amber band and target×1.30 trajectory
    # "At Risk" band so behaviour is unchanged until an admin edits them.
    "rag_amber_tolerance": 0.20,
    "trajectory_atrisk_tolerance": 0.30,
}

# (legacy _DEFAULT_RULES removed — cleaning rules now come from RULE_REGISTRY,
#  surfaced via /settings/rules and persisted under cleaning.enabled_rules.)


def _get_setting(key: str, default, db) -> dict:
    from .db.models import AppSetting

    row = db.query(AppSetting).filter_by(key=key).first()
    if row:
        return _json.loads(row.value)
    return default


def _load_kpi_targets(db) -> dict:
    """Current KPI targets for analytics: official defaults overlaid with any
    admin override stored under 'kpi.targets'. Returns {npan, who} dicts ready
    to pass into compute_kpi_dashboard / compute_trajectory_narratives."""
    defaults = official_targets()
    stored = _get_setting("kpi.targets", {}, db)
    return {
        "npan": {**defaults["npan"], **(stored.get("npan") or {})},
        "who": {**defaults["who"], **(stored.get("who") or {})},
    }


def _rag_tolerances(db) -> tuple[float, float]:
    """(amber_tolerance, atrisk_tolerance) from saved thresholds, falling back
    to the historical defaults when an older threshold.all has no such keys."""
    thr = _get_setting("threshold.all", _DEFAULT_THRESHOLDS, db)
    return (
        float(
            thr.get("rag_amber_tolerance", _DEFAULT_THRESHOLDS["rag_amber_tolerance"])
        ),
        float(
            thr.get(
                "trajectory_atrisk_tolerance",
                _DEFAULT_THRESHOLDS["trajectory_atrisk_tolerance"],
            )
        ),
    )


def _set_setting(key: str, value, db) -> None:
    from .db.models import AppSetting

    existing = db.query(AppSetting).filter_by(key=key).first()
    if existing:
        existing.value = _json.dumps(value)
    else:
        db.add(AppSetting(key=key, value=_json.dumps(value)))
    db.commit()


@app.get("/settings/thresholds")
def get_thresholds(db=Depends(get_db)):
    return _get_setting("threshold.all", _DEFAULT_THRESHOLDS, db)


@app.post("/settings/thresholds")
def post_thresholds(updates: dict, db=Depends(get_db)):
    current = _get_setting("threshold.all", _DEFAULT_THRESHOLDS, db)
    current.update(updates)
    _set_setting("threshold.all", current, db)
    _log_audit(
        action="settings.thresholds",
        detail=",".join(f"{k}={v}" for k, v in updates.items()),
    )
    return current


def _load_rule_state(db) -> dict:
    """{code: enabled} for every drop registry rule (B3). Default all-on; locked rules
    forced on; overlaid with the stored cleaning.enabled_rules selection."""
    stored = _get_setting("cleaning.enabled_rules", {}, db) or {}
    state = {}
    for code, meta in RULE_REGISTRY.items():
        if meta.get("locked"):
            state[code] = True
        else:
            v = stored.get(code)
            state[code] = True if v is None else bool(v)
    return state


def _load_review_rule_state(db) -> dict:
    """{code: enabled} for every active review rule. Default all-on; overlaid with
    stored cleaning.enabled_rules selection. Deferred rules (those with no source_types)
    are excluded."""
    stored = _get_setting("cleaning.enabled_rules", {}, db) or {}
    state = {}
    active_codes = {
        c for codes in REVIEW_EVALUATED_RULES.values() for c in codes
    }
    for code in REVIEW_RULE_REGISTRY:
        if code not in active_codes:
            continue  # skip deferred rules
        v = stored.get(code)
        state[code] = True if v is None else bool(v)
    return state


def _effective_enabled_rules(raw_body_rules, db):
    """Build the effective enabled-rules set passed to the cleaner for a run (D2).

    = (drop selection from the request, or the persisted drop state if the request
    omits it) UNION the persisted ENABLED review codes. The _REVIEW_MANAGED_SENTINEL
    is added whenever review rules are managed, so an all-disabled review selection
    is honoured (no review_* codes would otherwise read as "unmanaged -> all on").
    Returns None for the legacy all-on default (nothing passed AND reviews unmanaged).
    """
    try:
        persisted_stored = _get_setting("cleaning.enabled_rules", {}, db) or {}
    except Exception:
        persisted_stored = {}
    review_managed = any(k.startswith("review_") for k in persisted_stored)
    if raw_body_rules is None and not review_managed:
        return None  # legacy all-on: drops and reviews both default ON
    # Drop selection: from the request if given, else the persisted drop state.
    # Strip any review_* the body carried — the review portion is owned SOLELY by
    # persisted Settings, so a request can never re-enable a disabled review rule.
    if raw_body_rules is not None:
        base = {c for c in raw_body_rules if not str(c).startswith("review_")}
    else:
        try:
            base = {c for c, en in _load_rule_state(db).items() if en}
        except Exception:
            base = set()
    if not review_managed:
        return base  # reviews unmanaged -> cleaner leaves them all ON by default
    # Reviews managed: the resolved review selection (every active rule default-on,
    # overlaid with the stored disables) is authoritative — NOT the sparse store.
    # Sentinel marks "reviews managed" so an all-disabled selection is honoured
    # rather than read as unmanaged (which would default them all back ON).
    try:
        review_enabled = {c for c, en in _load_review_rule_state(db).items() if en}
    except Exception:
        review_enabled = set()
    return base | review_enabled | {_REVIEW_MANAGED_SENTINEL}


def _rule_source_types(code: str) -> list[str]:
    """Which source schemas actually run this drop rule (derived from EVALUATED_RULES)."""
    return [dt for dt, codes in EVALUATED_RULES.items() if code in codes]


def _review_rule_source_types(code: str) -> list[str]:
    """Which source schemas actually run this review rule (derived from REVIEW_EVALUATED_RULES)."""
    return [dt for dt, codes in REVIEW_EVALUATED_RULES.items() if code in codes]


def _rules_view(db) -> dict:
    drop_state = _load_rule_state(db)
    review_state = _load_review_rule_state(db)
    drop_rules = [
        {
            "code": c, **m,
            "enabled": drop_state[c],
            "source_types": _rule_source_types(c),
            "kind": "drop",
        }
        for c, m in RULE_REGISTRY.items()
    ]
    review_rules = [
        {
            "code": c, **REVIEW_RULE_REGISTRY[c],
            "enabled": review_state[c],
            "source_types": _review_rule_source_types(c),
            "locked": False,
            "kind": "review",
        }
        for c in review_state  # already excludes deferred
    ]
    return {"rules": drop_rules + review_rules}


@app.get("/settings/rules")
def get_rules(db=Depends(get_db)):
    """Registry-driven cleaning-rule state (B3). Replaces the inert rules.all set
    the cleaners never read — the SAME codes now drive the pipeline."""
    return _rules_view(db)


@app.post("/settings/rules/toggle")
def toggle_rule(body: dict, db=Depends(get_db)):
    rule = body.get("rule")
    enabled = bool(body.get("enabled", True))
    # Accept both drop rules and active review rules; reject unknown codes.
    _active_review_codes = {
        c for codes in REVIEW_EVALUATED_RULES.values() for c in codes
    }
    if rule not in RULE_REGISTRY and rule not in _active_review_codes:
        raise HTTPException(status_code=404, detail=f"Rule '{rule}' not found")
    if rule in LOCKED_RULES:
        raise HTTPException(
            status_code=400, detail=f"Rule '{rule}' is locked and always runs"
        )
    stored = _get_setting("cleaning.enabled_rules", {}, db) or {}
    stored[rule] = enabled
    _set_setting("cleaning.enabled_rules", stored, db)
    _log_audit(
        action="settings.rule_toggle", detail=f"{rule}={'on' if enabled else 'off'}"
    )
    return _rules_view(db)


def _load_target_year(db) -> int | None:
    """Admin-configured forecast target year (Settings → KPI Targets). None →
    auto-derive from the data (latest year + default horizon). Kept as its own
    setting key, not folded into kpi.targets, to avoid that round-trip's
    history of fragility."""
    v = _get_setting("kpi.target_year", None, db)
    try:
        return int(v) if v is not None else None
    except (TypeError, ValueError):
        return None


def _kpi_targets_view(db) -> dict:
    """GET-shaped view: current values, official defaults, and per-set source
    marker (official label vs 'custom' when a value differs from the default)."""
    defaults = official_targets()
    current = _load_kpi_targets(db)
    return {
        "current": current,
        "defaults": defaults,
        "target_year": _load_target_year(db),
        "source": {
            "npan": "custom"
            if current["npan"] != defaults["npan"]
            else "npan_2021_2025",
            "who": "custom" if current["who"] != defaults["who"] else "who_2025",
        },
    }


@app.get("/settings/kpi-targets")
def get_kpi_targets(db=Depends(get_db)):
    return _kpi_targets_view(db)


@app.post("/settings/kpi-targets")
def post_kpi_targets(body: dict, user=Depends(require_admin), db=Depends(get_db)):
    """Admin-only. Validates each target is numeric and in 0–100, merges over any
    existing override (so editing one set never drops the other), and audits who
    changed what. WHO and NPAN are both editable; labels are never touched."""
    defaults = official_targets()
    stored = _get_setting("kpi.targets", {}, db)
    out = {"npan": dict(stored.get("npan") or {}), "who": dict(stored.get("who") or {})}
    for grp in ("npan", "who"):
        incoming = body.get(grp)
        if incoming is None:
            continue
        if not isinstance(incoming, dict):
            raise HTTPException(status_code=422, detail=f"'{grp}' must be an object")
        for k, v in incoming.items():
            if k not in defaults[grp]:
                raise HTTPException(
                    status_code=422, detail=f"Unknown KPI key '{k}' in {grp}"
                )
            try:
                fv = float(v)
            except (TypeError, ValueError):
                raise HTTPException(
                    status_code=422, detail=f"{grp}.{k} must be numeric"
                )
            if not (0.0 <= fv <= 100.0):
                raise HTTPException(
                    status_code=422, detail=f"{grp}.{k} must be between 0 and 100"
                )
            out[grp][k] = fv
    # Optional forecast target year (separate key; null/empty clears → auto).
    if "target_year" in body:
        ty = body.get("target_year")
        if ty is None or ty == "":
            _set_setting("kpi.target_year", None, db)
        else:
            try:
                ty_int = int(ty)
            except (TypeError, ValueError):
                raise HTTPException(
                    status_code=422, detail="target_year must be an integer year"
                )
            if not (2020 <= ty_int <= 2100):
                raise HTTPException(
                    status_code=422, detail="target_year must be between 2020 and 2100"
                )
            _set_setting("kpi.target_year", ty_int, db)
    _set_setting("kpi.targets", out, db)
    _log_audit(
        action="settings.kpi_targets",
        detail=";".join(f"{g}.{k}={v}" for g, d in out.items() for k, v in d.items()),
        user_id=user.id,
    )
    return _kpi_targets_view(db)
